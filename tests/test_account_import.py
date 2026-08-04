"""账号导入：Excel → accounts（identity_status='imported'，待注册）。

这条链路以前根本不存在——只有导出没有导入，账号是手工塞进库的。

两条最容易出错、后果又最隐蔽的性质在这里钉死：

  1. 导入状态必须是 'imported'。补号流程 (_registerable_imported) 只认这个值；
     写成 'registered' 等于宣称 GitHub 已经开好了，账号既不会被注册、又因为没有
     GitHub 密码而登录不了，静静卡在列表里。
  2. 没有收码链接的账号虽然入库，却**领不走**（_hotmail_for_account 取不到收码数据
     就把它过滤掉）。导入接口必须把这批单独报出来，否则用户只会看到「导入成功」，
     然后困惑于补号流程一个都不碰。
"""

import openpyxl
import pytest

from src.models.account import AccountModel
from src.services import account_import

LINK = 'https://ruoanzhu.example/mail?e=a@x.com&p=1&h=2'


def _xlsx(tmp_path, header, rows, name='in.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(header, 1):
        ws.cell(row=1, column=i, value=h)
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    p = tmp_path / name
    wb.save(p)
    return str(p)


# ---------- 模版 ----------


def test_template_has_the_three_required_columns(tmp_path):
    p = account_import.generate_template(str(tmp_path / 't.xlsx'))
    ws = openpyxl.load_workbook(p).active
    header = [ws.cell(row=1, column=i).value for i in range(1, 4)]
    assert header == ['邮箱', '邮箱密码', '邮箱认证链接']


def test_generated_template_can_be_parsed_back(tmp_path):
    """模版自己必须导得进去——否则用户照着填完才发现格式不对。"""
    p = account_import.generate_template(str(tmp_path / 't.xlsx'))
    rows, errors = account_import.parse_excel(p)
    assert len(rows) == 1 and rows[0]['email'] == 'example001@hotmail.com'
    assert errors == []


# ---------- 解析 ----------


def test_parses_the_three_fields(tmp_path):
    p = _xlsx(tmp_path, ['邮箱', '邮箱密码', '邮箱认证链接'],
              [['a@x.com', 'pw1', LINK]])
    rows, errors = account_import.parse_excel(p)
    assert rows == [{'email': 'a@x.com', 'email_password': 'pw1',
                     'email_verify_link': LINK}]
    assert errors == []


@pytest.mark.parametrize('header', [
    ['email', 'password', 'link'],
    ['Email', 'Email_Password', 'Verify_Link'],
    ['账号', '密码', '收信链接'],
    ['邮箱地址', 'pwd', '认证链接'],
])
def test_column_names_are_matched_loosely(tmp_path, header):
    """列名放宽匹配：这张表多半是从邮箱供应商拷来的，写法五花八门。
    为一个列名对不上就整批导不进去太蠢。"""
    p = _xlsx(tmp_path, header, [['a@x.com', 'pw', LINK]])
    rows, _ = account_import.parse_excel(p)
    assert rows and rows[0]['email'] == 'a@x.com'
    assert rows[0]['email_password'] == 'pw'
    assert rows[0]['email_verify_link'] == LINK


def test_missing_email_column_is_a_hard_error(tmp_path):
    p = _xlsx(tmp_path, ['密码', '链接'], [['pw', LINK]])
    rows, errors = account_import.parse_excel(p)
    assert rows == []
    assert any('邮箱列' in e for e in errors)


def test_blank_rows_are_skipped_silently(tmp_path):
    """尾部空行太常见，报出来只是噪声。"""
    p = _xlsx(tmp_path, ['邮箱', '邮箱密码', '邮箱认证链接'],
              [['a@x.com', 'pw', LINK], [None, None, None], ['', '', '']])
    rows, errors = account_import.parse_excel(p)
    assert len(rows) == 1
    assert errors == []


def test_bad_and_duplicate_rows_are_reported_but_do_not_abort(tmp_path):
    """能导多少导多少。几百行里有两行不对就整批退回，只会逼用户来回试。"""
    p = _xlsx(tmp_path, ['邮箱', '邮箱密码', '邮箱认证链接'], [
        ['a@x.com', 'pw', LINK],
        ['not-an-email', 'pw', LINK],
        ['a@x.com', 'pw2', LINK],
        ['b@x.com', '', LINK],
    ])
    rows, errors = account_import.parse_excel(p)
    assert [r['email'] for r in rows] == ['a@x.com', 'b@x.com']
    assert len(errors) == 2


def test_optional_fields_may_be_empty(tmp_path):
    p = _xlsx(tmp_path, ['邮箱', '邮箱密码', '邮箱认证链接'], [['a@x.com', None, None]])
    rows, _ = account_import.parse_excel(p)
    assert rows[0]['email_password'] == '' and rows[0]['email_verify_link'] == ''


# ---------- 入库 ----------


def test_imported_accounts_land_in_the_pending_registration_state(db):
    """必须是 'imported' —— 补号流程只认这个值。"""
    am = AccountModel(db)
    account_import.import_rows(am, [
        {'email': 'a@x.com', 'email_password': 'pw', 'email_verify_link': LINK}])

    row = db.fetchone("SELECT * FROM accounts WHERE email='a@x.com'")
    assert row['identity_status'] == 'imported'
    assert row['email_password'] == 'pw'
    assert row['email_verify_link'] == LINK


def test_accounts_without_a_link_are_reported_separately(db):
    """没有收码链接的账号入了库也领不走，必须单独点出来。"""
    am = AccountModel(db)
    stat = account_import.import_rows(am, [
        {'email': 'a@x.com', 'email_password': 'pw', 'email_verify_link': LINK},
        {'email': 'b@x.com', 'email_password': 'pw', 'email_verify_link': ''},
    ])
    assert stat['imported'] == 2
    assert stat['no_link'] == ['b@x.com']


def test_reimport_does_not_wipe_existing_values(db):
    """重复导入同一份表要安全：已有的密码/链接不能被空值覆盖。"""
    am = AccountModel(db)
    account_import.import_rows(am, [
        {'email': 'a@x.com', 'email_password': 'pw', 'email_verify_link': LINK}])
    account_import.import_rows(am, [
        {'email': 'a@x.com', 'email_password': '', 'email_verify_link': ''}])

    row = db.fetchone("SELECT * FROM accounts WHERE email='a@x.com'")
    assert row['email_password'] == 'pw'
    assert row['email_verify_link'] == LINK


def test_imported_accounts_are_picked_up_for_registration(db):
    """端到端：导进来的账号确实会被补号流程领走。

    这才是导入功能的意义所在——状态对、收码数据齐，两者缺一就领不走。
    """
    from src.web.app import AppState

    am = AccountModel(db)
    account_import.import_rows(am, [
        {'email': 'has@x.com', 'email_password': 'pw', 'email_verify_link': LINK},
        {'email': 'none@x.com', 'email_password': 'pw', 'email_verify_link': ''},
    ])

    st = AppState.__new__(AppState)
    st._hotmail_map = None
    st._hooked_print = lambda *a, **k: None
    got = [a['email'] for a in am.get_all(order_desc=False)
           if (a.get('identity_status') or '') == 'imported' and st._hotmail_for_account(a)]

    assert got == ['has@x.com'], '缺收码链接的账号不该被领走'
