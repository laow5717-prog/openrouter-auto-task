#!/usr/bin/env python3
"""合并底料目录中的 Excel 文件，按照 credit_cards_template 的格式"""

import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter

def get_template_columns():
    """获取模版文件的列结构"""
    template_file = Path('credit_cards_template.xlsx')
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active

    # 获取第一行的列名
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    return headers

def extract_mapped_columns(source_ws, template_columns):
    """从源 sheet 中提取对应模版列的数据

    映射关系：
    - first_name → 名字
    - last_name → 姓氏
    - country → 账单国家
    - address → 账单地址1
    - address2 → 账单地址2
    - city → 账单城市
    - state → 账单省
    - zip → 账单邮编
    - card_number → 卡号
    - expiry_month/year → 有效期 (MM/YY格式)
    - cvc → CVV
    - company → (暂无映射)
    """

    # 从源文件获取列索引
    source_headers = {}
    for cell in source_ws[1]:
        if cell.value:
            source_headers[cell.value] = cell.column

    # 列映射关系
    mapping = {
        'card_number': '卡号',
        'expiry_month': '有效期',  # 需要特殊处理
        'expiry_year': '有效期',   # 需要特殊处理
        'cvc': 'CVV',
        'first_name': '名字',
        'last_name': '姓氏',
        'country': '账单国家',
        'address': '账单地址1',
        'address2': '账单地址2',
        'city': '账单城市',
        'state': '账单省',
        'zip': '账单邮编',
        'company': None,  # 暂无映射
    }

    rows_data = []
    for row_idx, row in enumerate(source_ws.iter_rows(min_row=2, values_only=False), 2):
        if not any(cell.value for cell in row):
            continue

        row_data = []
        for template_col in template_columns:
            source_col_name = mapping.get(template_col)

            if source_col_name is None:
                # company 列暂无映射
                row_data.append(None)
            elif source_col_name in source_headers:
                col_idx = source_headers[source_col_name]
                cell_value = source_ws.cell(row=row_idx, column=col_idx).value

                # 有效期特殊处理：从 "MM/YY" 格式提取
                if template_col in ['expiry_month', 'expiry_year']:
                    if cell_value:
                        try:
                            # 格式是 MM/YY (例如 03/31)
                            parts = str(cell_value).split('/')
                            if len(parts) == 2:
                                if template_col == 'expiry_month':
                                    row_data.append(parts[0].zfill(2))  # MM
                                else:  # expiry_year
                                    yy = parts[1].zfill(2)
                                    # 如果YY < 70，认为是 20xx，否则是 19xx
                                    year = int(yy)
                                    if year < 70:
                                        row_data.append('20' + yy)
                                    else:
                                        row_data.append('19' + yy)
                            else:
                                row_data.append(None)
                        except:
                            row_data.append(None)
                    else:
                        row_data.append(None)
                else:
                    row_data.append(cell_value)
            else:
                row_data.append(None)

        rows_data.append(row_data)

    return rows_data

def merge_files():
    """合并所有底料文件"""
    base_dir = Path('底料')
    template_file = Path('credit_cards_template.xlsx')
    output_file = Path('merged_credit_cards.xlsx')

    if not base_dir.exists():
        print(f"❌ 底料目录不存在")
        return

    if not template_file.exists():
        print(f"❌ 模版文件不存在")
        return

    # 获取模版列
    template_columns = get_template_columns()
    print(f"✓ 模版列: {template_columns}")

    # 创建输出工作簿
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = 'CreditCards'

    # 写入模版列标题
    for col_idx, header in enumerate(template_columns, 1):
        output_ws.cell(row=1, column=col_idx, value=header)

    # 收集所有底料文件的数据
    all_rows = []
    xlsx_files = sorted(base_dir.glob('*.xlsx'))

    print(f"\n找到 {len(xlsx_files)} 个底料文件")

    for xlsx_file in xlsx_files:
        print(f"\n📄 处理: {xlsx_file.name}")
        try:
            source_wb = openpyxl.load_workbook(xlsx_file)
            source_ws = source_wb.active

            # 提取数据
            rows = extract_mapped_columns(source_ws, template_columns)
            all_rows.extend(rows)
            print(f"  ✓ 提取 {len(rows)} 行数据")
        except Exception as e:
            print(f"  ❌ 错误: {e}")

    # 写入输出文件
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            output_ws.cell(row=row_idx, column=col_idx, value=value)

    output_wb.save(output_file)
    print(f"\n✓ 合并完成！")
    print(f"  输出文件: {output_file}")
    print(f"  总计行数: {len(all_rows)}")

if __name__ == '__main__':
    merge_files()
