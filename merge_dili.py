#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""将 底料/ 目录下所有 xlsx 按 credit_cards_template.xlsx 结构合并。

用法（在项目根目录）:
    .venv/bin/python3 merge_dili.py
输出: 底料/merged_credit_cards.xlsx（覆盖前自动备份为 .bak-时间戳）
额外: 运行时会报告每个 sheet 中“未被字段映射命中的列”，用于发现漏映射。
"""
import glob, os, datetime, re, shutil
from collections import OrderedDict, Counter
import openpyxl

OUT_PATH = '底料/merged_credit_cards.xlsx'

TEMPLATE_COLUMNS = [
    "card_number", "expiry_month", "expiry_year", "cvc",
    "first_name", "last_name", "country", "address",
    "address2", "city", "state", "zip", "company",
]
CRITICAL = ["card_number", "expiry_month", "expiry_year", "cvc"]  # 缺则丢弃
IMPORT_REQUIRED = ["card_number", "expiry_month", "expiry_year", "cvc",
                   "first_name", "last_name", "country", "address", "city", "state", "zip"]

US_STATES = {
    'AL':'Alabama','AK':'Alaska','AZ':'Arizona','AR':'Arkansas','CA':'California',
    'CO':'Colorado','CT':'Connecticut','DE':'Delaware','FL':'Florida','GA':'Georgia',
    'HI':'Hawaii','ID':'Idaho','IL':'Illinois','IN':'Indiana','IA':'Iowa','KS':'Kansas',
    'KY':'Kentucky','LA':'Louisiana','ME':'Maine','MD':'Maryland','MA':'Massachusetts',
    'MI':'Michigan','MN':'Minnesota','MS':'Mississippi','MO':'Missouri','MT':'Montana',
    'NE':'Nebraska','NV':'Nevada','NH':'New Hampshire','NJ':'New Jersey','NM':'New Mexico',
    'NY':'New York','NC':'North Carolina','ND':'North Dakota','OH':'Ohio','OK':'Oklahoma',
    'OR':'Oregon','PA':'Pennsylvania','RI':'Rhode Island','SC':'South Carolina',
    'SD':'South Dakota','TN':'Tennessee','TX':'Texas','UT':'Utah','VT':'Vermont',
    'VA':'Virginia','WA':'Washington','WV':'West Virginia','WI':'Wisconsin','WY':'Wyoming',
    'DC':'District of Columbia','PR':'Puerto Rico',
}
STATE_FULL = {v.lower(): v for v in US_STATES.values()}

# header 同义词 -> 模版字段（按名匹配，兼容各文件不同列位置）
SYN = {
    'card_number': ['卡号'],
    'cvc': ['cvv'],
    'expiry_raw': ['有效期'],
    'expiry_month_col': ['有效期月份'],   # 7.9MK 把月份单列存放
    'expiry_year_hint': ['有效期年份', '有效期份', '年份'],
    'first_name': ['名字', '名'],
    'last_name': ['姓氏', '姓'],
    'holder': ['持卡人'],
    'country': ['账单国家', '国家', '国家名称'],   # 排除“卡片国家”(发卡国)
    'zip': ['账单邮编', '邮编'],
    'state': ['账单省', '州'],
    'city': ['账单城市', '城市'],
    'address': ['账单地址1', '地址1'],
    'address2': ['账单地址2', '地址2'],
    'company': ['银行', '发卡行'],
}

def norm_country(v):
    s = (v or '').strip()
    if not s:
        return 'United States'
    low = s.lower()
    if low in ('us', 'usa', 'united states', 'united states of america', '美国', 'unitedstates'):
        return 'United States'
    return s

def norm_state(v):
    s = (v or '').strip()
    if not s:
        return ''
    up = s.upper()
    if len(s) == 2 and up in US_STATES:
        return US_STATES[up]
    if s.lower() in STATE_FULL:
        return STATE_FULL[s.lower()]
    return s  # 未知，原样

def split_name(holder):
    s = (holder or '').strip()
    if not s:
        return '', ''
    parts = s.split()
    if len(parts) == 1:
        return parts[0], parts[0]
    return parts[0], ' '.join(parts[1:])

def parse_expiry(raw, year_hint=None):
    """返回 (month2, year4) 或 ('','')"""
    yhint = None
    if year_hint is not None:
        m = re.search(r'(\d{4})', str(year_hint))
        if m:
            yhint = m.group(1)
        else:
            m2 = re.search(r'(\d{2})', str(year_hint))
            if m2:
                yhint = '20' + m2.group(1)
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        # 7.10电商: 有效期存成 Excel 日期序列号(如 46382), 先还原为日期再按下方规则解析
        try:
            raw = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(raw))
        except (ValueError, OverflowError):
            return ('', yhint) if yhint else ('', '')
    if isinstance(raw, datetime.datetime):
        # 7.8电商: Excel 把 MM/YY(YY<=31且合法) 误存成日期, day 即年份后两位
        return f"{raw.month:02d}", f"20{raw.day:02d}"
    s = (raw or '').strip() if isinstance(raw, str) else ''
    if s:
        if '/' in s:
            a, b = (s.split('/', 1) + [''])[:2]
            a, b = a.strip(), b.strip()
            month = a.zfill(2) if a.isdigit() else ''
            if b.isdigit():
                year = b if len(b) == 4 else ('20' + b.zfill(2))
            else:
                year = yhint or ''
            return month, year
        if s.isdigit():
            if len(s) in (1, 2):
                return s.zfill(2), (yhint or '')
            if len(s) == 4:
                return s[:2], '20' + s[2:]
    return ('', yhint) if yhint else ('', '')

def build_colmap(header):
    hdr = [str(c).strip().lower() if c is not None else '' for c in header]
    cmap = {}
    for field, syns in SYN.items():
        for i, h in enumerate(hdr):
            if h in [x.lower() for x in syns]:
                cmap[field] = i
                break
    return cmap

def rows_from_sheet(ws, source_label):
    """返回 (rows, unmapped_cols)。unmapped_cols 为未命中任何字段映射的非空列名。"""
    data = list(ws.iter_rows(min_row=1, values_only=True))
    if len(data) < 2:
        return [], []
    cmap = build_colmap(data[0])
    if 'card_number' not in cmap:
        return [], []
    mapped_idx = set(cmap.values())
    unmapped = [str(h).strip() for i, h in enumerate(data[0])
                if i not in mapped_idx and h is not None and str(h).strip()]
    out = []
    for row in data[1:]:
        def g(f):
            i = cmap.get(f)
            if i is None or i >= len(row):
                return None
            return row[i]
        card = g('card_number')
        card = str(card).strip().replace(' ', '') if card is not None else ''
        if not card or not card.isdigit():
            continue
        first = str(g('first_name')).strip() if g('first_name') is not None else ''
        last = str(g('last_name')).strip() if g('last_name') is not None else ''
        # “持卡人”优先：MK 系列的“姓/名”两列语义与中文相反
        # （姓=Nate 名=Morgado 持卡人=Nate Morgado），直译会把 first/last 写反。
        # 持卡人能拆出两段时以它为准；否则缺哪个补哪个。
        hf, hl = split_name(g('holder')) if g('holder') is not None else ('', '')
        if hf and hl and hf != hl:
            first, last = hf, hl
        else:
            if not first:
                first = hf
            if not last:
                last = hl
        month, year = parse_expiry(g('expiry_raw'), g('expiry_year_hint'))
        # 若无“有效期”合并列、而是“有效期月份”单列（如 7.9MK），用它补月份
        if not month:
            dm = g('expiry_month_col')
            if dm is not None:
                dm_s = str(dm).strip()
                if dm_s.endswith('.0'):
                    dm_s = dm_s[:-2]
                if dm_s.isdigit():
                    month = dm_s.zfill(2)
        cvc = str(g('cvc')).strip() if g('cvc') is not None else ''
        if cvc.endswith('.0'):
            cvc = cvc[:-2]
        def sv(f):
            v = g(f)
            return str(v).strip() if v is not None else ''
        out.append({
            'card_number': card, 'expiry_month': month, 'expiry_year': year, 'cvc': cvc,
            'first_name': first, 'last_name': last, 'country': norm_country(g('country')),
            'address': sv('address'), 'address2': sv('address2'), 'city': sv('city'),
            'state': norm_state(g('state')), 'zip': sv('zip'), 'company': sv('company'),
            '_src': source_label,
        })
    return out, unmapped

def collect_all():
    files = sorted(glob.glob('底料/*.xlsx'))
    all_rows, per_src, unmapped_report = [], OrderedDict(), OrderedDict()
    seen_mao = False
    for f in files:
        base = os.path.basename(f)
        # 跳过合并产物本身，避免把上次输出当源读回
        if base.startswith('合并') or base.startswith('merged'):
            continue
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        if '猫头鹰260' in base:
            if seen_mao:
                per_src[base + ' (重复文件,跳过)'] = 0
                wb.close(); continue
            seen_mao = True
        for ws in wb.worksheets:
            label = f"{base}::{ws.title}"
            rs, unmapped = rows_from_sheet(ws, label)
            if rs:
                per_src[label] = len(rs)
                all_rows.extend(rs)
                if unmapped:
                    unmapped_report[label] = unmapped
        wb.close()
    return all_rows, per_src, unmapped_report

def merge(all_rows):
    merged = OrderedDict()
    for r in all_rows:
        c = r['card_number']
        if c not in merged:
            merged[c] = {k: r[k] for k in TEMPLATE_COLUMNS}
        else:
            tgt = merged[c]
            for k in TEMPLATE_COLUMNS:
                if not tgt[k] and r[k]:
                    tgt[k] = r[k]
    return merged

def main():
    all_rows, per_src, unmapped_report = collect_all()
    merged = merge(all_rows)
    kept, dropped = [], []
    for c, rec in merged.items():
        miss = [k for k in CRITICAL if not rec[k]]
        (dropped if miss else kept).append((c, miss) if miss else rec)

    out_path = OUT_PATH
    # 覆盖前自动备份现有产物
    if os.path.exists(out_path):
        stamp = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
        bak = f"{out_path}.bak-{stamp}"
        shutil.copy2(out_path, bak)
        print(f"已备份原产物 -> {bak}")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'CreditCards'
    for j, col in enumerate(TEMPLATE_COLUMNS, 1):
        ws.cell(row=1, column=j, value=col)
    for i, rec in enumerate(kept, 2):
        for j, col in enumerate(TEMPLATE_COLUMNS, 1):
            ws.cell(row=i, column=j, value=rec[col])
    wb.save(out_path)

    fully_valid = sum(1 for r in kept if all(r[k] for k in IMPORT_REQUIRED))
    print("=== 各来源(sheet)有效行数 ===")
    for k, v in per_src.items():
        print(f"  {k}: {v}")
    if unmapped_report:
        print("\n=== 未映射的列（未纳入模版，请确认是否需要补映射）===")
        for label, cols in unmapped_report.items():
            print(f"  {label}: {cols}")
    print(f"\n原始行合计: {len(all_rows)}")
    print(f"去重后唯一卡号: {len(merged)}")
    print(f"丢弃(缺卡号/有效期/CVV): {len(dropped)}")
    print(f"写入输出: {len(kept)} 行 -> {out_path}")
    print(f"其中完全满足导入必填(11项)的: {fully_valid} 行")
    miss_counter = Counter()
    for r in kept:
        for k in IMPORT_REQUIRED:
            if not r[k]:
                miss_counter[k] += 1
    if miss_counter:
        print("\n导入必填的缺项分布(这些行上传会被导入器拒绝):")
        for k, v in miss_counter.most_common():
            print(f"  {k}: 缺 {v} 行")
    dc = Counter()
    for _, miss in dropped:
        dc[','.join(miss)] += 1
    if dc:
        print("\n丢弃原因分布:")
        for k, v in dc.most_common():
            print(f"  缺[{k}]: {v} 行")

    # 按来源统计关键字段缺失（去重前），用于定位是哪个文件缺 expiry / name
    diag_fields = ['first_name', 'last_name', 'expiry_month', 'expiry_year',
                   'cvc', 'city', 'state', 'zip', 'address', 'country']
    src_stat = OrderedDict()
    for r in all_rows:
        d = src_stat.setdefault(r['_src'], Counter())
        d['__total'] += 1
        for f in diag_fields:
            if not r[f]:
                d[f] += 1
    print("\n=== 按来源字段缺失诊断（去重前，仅列出有缺失的字段）===")
    for s, d in src_stat.items():
        miss = {f: d[f] for f in diag_fields if d[f]}
        print(f"  {s} (共{d['__total']}行): {miss if miss else '无缺失'}")

if __name__ == '__main__':
    main()
