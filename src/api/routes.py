"""
Flask API 路由
所有 /api/* 端点
"""

import os
import io
import json
from flask import Blueprint, jsonify, request, send_from_directory, send_file

from src.config import cfg, get_data_dir, RechargeConfig
from src.services import card as card_service
from src.services import registration
from src.models.card_pool import CardPoolModel
from src.utils import is_card_expired, is_identity_terminal, is_platform_terminal

api = Blueprint('api', __name__)


def get_app_state():
    """默认平台的运行上下文。

    ⚠️ 多平台并发后，「AppState」已经不是全局唯一的了——每个平台一个。
    凡是与**某次运行**有关的操作（启动/停止/看状态/看日志），都必须用
    `get_ctx()` 按平台取，用这个函数会永远操作默认平台。

    保留它是因为有一批接口只碰共享资源（models、db、registry），
    用哪个 ctx 都一样；那些地方继续用它没问题。
    """
    from flask import current_app
    return current_app.config['APP_STATE']


def get_contexts():
    """全部平台的运行上下文，{slug: ctx}。"""
    from flask import current_app
    return current_app.config.get('RUN_CONTEXTS') or {
        current_app.config['APP_STATE'].platform: current_app.config['APP_STATE']
    }


def get_ctx(platform=None):
    """取某个平台的运行上下文。platform 为空时用请求里的平台参数。

    找不到对应平台时回落到默认 ctx 而不是报错：平台列表由代码里的适配器注册表
    决定，传了未知 slug 属于调用方的问题，但让整个接口 500 更难查。
    """
    ctxs = get_contexts()
    slug = platform or _req_platform()
    return ctxs.get(slug) or get_app_state()


def get_db():
    from flask import current_app
    return current_app.config['DB']


def get_models():
    from flask import current_app
    return current_app.config['MODELS']


def _req_platform(required=False):
    """取本次请求的目标平台 slug。

    读接口（账号/卡池列表）缺省时回落到**默认平台**，让不带参数的老调用仍能工作。
    写接口与卡池类接口应传 required=True——猜错平台会返回或写入混合数据，
    那比直接报错糟糕得多。

    兜底值曾经是「AppState.platform」，也就是**上一次运行的那个平台**。单平台时
    它恰好等于用户正在看的平台，所以一直没出事；两个平台同时跑之后，这个值会被
    后启动的那个覆盖，于是不带参数的请求会随机地读到另一个平台的数据。
    现在改成固定的默认平台常量——猜得稳定，总比猜得随机好。
    """
    value = (request.args.get('platform')
             or (request.get_json(silent=True) or {}).get('platform')
             or '')
    value = value.strip()
    if value:
        return value
    if required:
        return None
    from src.web.app import AppState
    return AppState.DEFAULT_PLATFORM


@api.route('/api/settings/adspower', methods=['GET'])
def get_adspower_settings():
    """AdsPower 的生效配置。api_key **明文**返回。

    明文是有意的：本机单人使用，同一个库里 GitHub 密码、邮箱密码本来就明文躺着，
    单给这一个字段打码挡不住任何真实威胁，却要引入「提交上来的是新 key 还是掩码」
    的判断——判错就把真 key 覆盖成掩码串，那才是真实的破坏。

    from_db 让界面能区分「这是我设的」还是「这是 config.yaml 的默认值」。
    """
    from src.models import settings as S
    models = get_models()
    eff = get_app_state().adspower_settings()
    stored = models['settings'].get_many(
        [S.KEY_ADSPOWER_ENABLED, S.KEY_ADSPOWER_API_KEY, S.KEY_ADSPOWER_BASE_URL])
    return jsonify({
        "enabled": eff['enabled'],
        "base_url": eff['base_url'],
        "api_key": eff['api_key'],
        "from_db": {
            "enabled": S.KEY_ADSPOWER_ENABLED in stored,
            "api_key": S.KEY_ADSPOWER_API_KEY in stored,
            "base_url": S.KEY_ADSPOWER_BASE_URL in stored,
        },
    })


@api.route('/api/settings/adspower', methods=['PUT'])
def save_adspower_settings():
    """保存 AdsPower 配置。只写传上来的字段，没传的保持原样。

    每个字段的三种取值语义：
      - 不传（字段缺席）→ 不动
      - 传空串          → **清除**覆盖值，回落 config.yaml 的默认值
      - 传值            → 覆盖
    「缺席」与「空串」必须区分：都当成清除的话，前端只想改开关也会把 key 抹掉。
    """
    from src.models import settings as S
    models = get_models()
    data = request.json or {}
    sm = models['settings']

    if 'enabled' in data:
        sm.set(S.KEY_ADSPOWER_ENABLED, '1' if data.get('enabled') else '0')

    if 'base_url' in data:
        url = (data.get('base_url') or '').strip()
        if url and not url.startswith(('http://', 'https://')):
            return jsonify({"error": "地址需以 http:// 或 https:// 开头"}), 400
        sm.set(S.KEY_ADSPOWER_BASE_URL, url or None)

    if 'api_key' in data:
        key = (data.get('api_key') or '').strip()
        sm.set(S.KEY_ADSPOWER_API_KEY, key or None)

    return get_adspower_settings()


@api.route('/api/settings/adspower/test', methods=['POST'])
def test_adspower_settings():
    """连通性自检：客户端可达吗、key 有效吗。

    没有它的话，key 填错只会在下一次跑任务时表现成一句「浏览器起不来」，
    与配置页隔着十万八千里，用户根本对不上因果。
    """
    state = get_app_state()
    s = state.adspower_settings()
    if not s['api_key']:
        return jsonify({"ok": False, "detail": "尚未配置 API Key"})

    # 复用**共享**的那个客户端，不新建。AdsPowerClient 的 _throttle 限流状态是
    # 实例级的（见 SharedResources 的 docstring）：多一个实例就等于多一倍请求速率，
    # 会撞 AdsPower 本地接口的频率限制——任务正在跑时点一下检测，可能把任务一起撞挂。
    # 2026-08-05 连点四次检测就复现过，第四次直接被拒、报成「API Key 不对」。
    #
    # 安全性由 _ensure_adspower 保证：它按 (api_key, base_url) 比对，配置变了会重建，
    # 所以拿到的客户端一定是当前生效配置的那个，不存在「验的是旧参数」。
    client, _pool = state._ensure_adspower()
    if client is None:
        # 开关关着时没有共享实例可用。此处只发一个请求，临时建一个可以接受——
        # 而且这条路径下不会有任务在跑（任务用的就是这个开关）。
        from src.services.adspower import AdsPowerClient
        client = AdsPowerClient(s['base_url'], s['api_key'])
    try:
        profiles = client.list_profiles(page_size=1)
        return jsonify({"ok": True,
                        "detail": f"连接正常，当前环境数可读（返回 {len(profiles)} 条样本）"})
    except Exception as e:
        msg = str(e)[:200]
        if 'Connection' in msg or 'refused' in msg or 'timed out' in msg.lower():
            hint = "连不上客户端——请确认 AdsPower 已启动，且「本地API」开关已打开"
        else:
            hint = "客户端有响应但请求被拒——多半是 API Key 不对"
        return jsonify({"ok": False, "detail": f"{hint}（{msg}）"})


@api.route('/api/platforms')
def list_platforms():
    """可用平台列表 + 当前选中平台。

    真值源是代码里的适配器注册表，不是数据库——平台随代码发布变化，不是运行时数据。
    """
    import src.platforms as platforms
    return jsonify({
        "data": platforms.describe_all(),
        # 「当前」的语义已经变了：多平台并发时没有唯一的当前平台。这里返回**默认**
        # 平台，仅供前端首次加载时选一个；真正在看哪个由前端 store 自己持有。
        "current": get_app_state().platform,
        "running": [slug for slug, c in get_contexts().items() if c.is_running],
    })


@api.route('/api/status')
def get_status():
    """全局状态 + 各 worker 概览。

    向后兼容：顶层字段全部保留原语义，workers 数组是新增的旁挂字段。
    串行运行时 workers 只有 W1，顶层 current_action 即 W1 的动作。

    多平台：顶层字段是**所请求平台**的状态（platform 参数，缺省用默认平台）。
    另外旁挂一个 `platforms` 汇总，让前端能看见**没在看的那个平台**是否在跑——
    否则它出问题时用户完全看不见。
    """
    state = get_ctx()
    models = get_models()

    total_inventory = models['account'].count()

    return jsonify({
        "platform": state.platform,
        "is_running": state.is_running,
        "current_action": state.current_action,
        "success": state.success_count,
        "fail": state.fail_count,
        "total_inventory": total_inventory,
        "logs": state.get_logs(int(request.args.get('log_index', 0))),
        "parallel_mode": state.parallel_mode,
        "workers": [
            {
                "id": w.worker_id,
                "current_action": w.current_action,
                "busy": w.busy,
                "log_seq": w.log_seq,
            }
            for w in state.active_workers()
        ],
        # 全平台概览。刻意只放很轻的字段——它每秒被轮询一次。
        "platforms": {
            slug: {
                "is_running": c.is_running,
                "current_action": c.current_action,
                "success": c.success_count,
                "fail": c.fail_count,
            }
            for slug, c in get_contexts().items()
        },
        "quota": state.quota.snapshot() if hasattr(state, 'quota') else None,
    })


@api.route('/api/workers/<worker_id>/logs')
def get_worker_logs(worker_id):
    """按 worker 增量拉取日志。

    index 传上次返回的 next_index；worker 不存在时回落到主 worker，
    保证前端在 worker 数变化时不会拿到 404。

    必须按平台取 ctx：两个平台各有一套同名的 W1..W4，取错 ctx 会拿到
    另一个平台那个 worker 的日志——而且因为有「回落主 worker」的兜底，
    不会 404，只会安静地给错数据。
    """
    state = get_ctx()
    worker = state.get_worker(worker_id)
    logs, next_index = worker.get_logs(int(request.args.get('index', 0)))
    return jsonify({
        "worker_id": worker.worker_id,
        "logs": logs,
        "next_index": next_index,
        "current_action": worker.current_action,
        "busy": worker.busy,
    })


@api.route('/api/start', methods=['POST'])
def start_task():
    state = get_ctx()
    if state.is_running:
        return jsonify({"error": f"{state.platform} 已有任务在运行"}), 400

    data = request.json or {}
    count = data.get('count', 1)
    card_info_list = data.get('card_info_list', None)
    login_password = data.get('login_password', None)
    max_bindable_cards = data.get('max_bindable_cards', 2)
    captcha_api_key = data.get('captcha_api_key', None)

    import threading
    threading.Thread(
        target=state.run_batch_task,
        args=(count, card_info_list, login_password, max_bindable_cards, captcha_api_key),
        daemon=True,
    ).start()

    return jsonify({"status": "started"})


@api.route('/api/stop', methods=['POST'])
def stop_task():
    """停止**某一个平台**的任务，不影响另一个平台（AC2）。"""
    state = get_ctx()
    if not state.is_running:
        return jsonify({"error": f"{state.platform} 没有正在运行的任务"}), 400

    state.force_stop()
    return jsonify({"status": "stopping", "platform": state.platform})


@api.route('/api/card/template')
def download_card_template():
    template_path = card_service.generate_template()
    directory = os.path.dirname(template_path)
    filename = os.path.basename(template_path)
    return send_from_directory(directory, filename, as_attachment=True)


@api.route('/api/card/history')
def get_card_history():
    models = get_models()

    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 20))
    status = request.args.get('status', '')
    keyword = request.args.get('keyword', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    records, total = models['card_binding'].get_all_paginated(
        page=page, page_size=page_size,
        status=status, keyword=keyword,
        date_from=date_from, date_to=date_to,
    )
    summary = models['card_binding'].get_global_summary()

    formatted = []
    for r in records:
        card = {}
        if r.get('card_data_json'):
            try:
                card = json.loads(r['card_data_json'])
            except (json.JSONDecodeError, TypeError):
                pass
        formatted.append({
            "task_id": r.get('task_id'),
            "card_number": card.get('number', ''),
            "card_holder": f"{card.get('first_name', '')} {card.get('last_name', '')}".strip(),
            "expiry_month": card.get('expiry_month', ''),
            "expiry_year": card.get('expiry_year', ''),
            "cvc": card.get('cvc', ''),
            "country": card.get('country', ''),
            "address": card.get('address', ''),
            "address2": card.get('address2', ''),
            "city": card.get('city', ''),
            "state": card.get('state', ''),
            "zip": card.get('zip', ''),
            "company": card.get('company', ''),
            "status": r['status'],
            "bound_to_email": r.get('bound_to_email') or '',
            "error": r.get('error') or '',
            "attempted_at": r.get('attempted_at') or '',
        })

    return jsonify({
        "data": formatted,
        "total": total,
        "page": page,
        "page_size": page_size,
        "summary": summary,
    })


@api.route('/api/card/history/cleanup', methods=['POST'])
def cleanup_card_history():
    models = get_models()
    # 保护**所有**平台正在跑的任务，不只当前这个——多平台并发时可能同时有两个
    # 批量任务在跑，只保护其中一个会把另一个的绑卡记录删掉。
    active = [c.current_card_task_id for c in get_contexts().values()
              if c.is_running and c.current_card_task_id]
    deleted = models['card_binding'].cleanup_stale_pending(active)
    return jsonify({"deleted": deleted})


@api.route('/api/card/history/export', methods=['POST'])
def export_card_history():
    models = get_models()
    data = request.json or {}

    status = data.get('status', '')
    keyword = data.get('keyword', '')
    date_from = data.get('date_from', '')
    date_to = data.get('date_to', '')

    records = models['card_binding'].get_all_filtered(
        status=status, keyword=keyword,
        date_from=date_from, date_to=date_to,
    )

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "绑卡记录"

    headers = ["序号", "卡号", "状态", "绑定账号", "错误信息", "处理时间", "任务批次",
               "完整卡号", "有效期", "CVC", "持卡人",
               "国家", "地址1", "地址2", "城市", "州/省", "邮编", "公司"]
    header_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    header_font = Font(bold=True, size=11)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, r in enumerate(records, 2):
        card = {}
        if r.get('card_data_json'):
            try:
                card = json.loads(r['card_data_json'])
            except (json.JSONDecodeError, TypeError):
                pass

        ws.cell(row=row_idx, column=1, value=r['id'])
        # card_bindings.card_display 存的是后四位，优先用 card_data_json 里的完整卡号
        ws.cell(row=row_idx, column=2, value=card.get('number') or r['card_display'])
        ws.cell(row=row_idx, column=3, value=r['status'])
        ws.cell(row=row_idx, column=4, value=r.get('bound_to_email') or '')
        ws.cell(row=row_idx, column=5, value=r.get('error') or '')
        ws.cell(row=row_idx, column=6, value=r.get('attempted_at') or '')
        ws.cell(row=row_idx, column=7, value=r.get('task_id') or '')
        ws.cell(row=row_idx, column=8, value=card.get('number', ''))
        expiry = f"{card.get('expiry_month', '')}/{card.get('expiry_year', '')}" if card.get('expiry_month') else ''
        ws.cell(row=row_idx, column=9, value=expiry)
        ws.cell(row=row_idx, column=10, value=card.get('cvc', ''))
        ws.cell(row=row_idx, column=11, value=f"{card.get('first_name', '')} {card.get('last_name', '')}".strip())
        ws.cell(row=row_idx, column=12, value=card.get('country', ''))
        ws.cell(row=row_idx, column=13, value=card.get('address', ''))
        ws.cell(row=row_idx, column=14, value=card.get('address2', ''))
        ws.cell(row=row_idx, column=15, value=card.get('city', ''))
        ws.cell(row=row_idx, column=16, value=card.get('state', ''))
        ws.cell(row=row_idx, column=17, value=card.get('zip', ''))
        ws.cell(row=row_idx, column=18, value=card.get('company', ''))

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='card_history_export.xlsx',
    )


@api.route('/api/accounts')
def get_accounts():
    models = get_models()

    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 20))
    keyword = request.args.get('keyword', '')
    identity_status = request.args.get('identity_status', '')
    platform_status = request.args.get('platform_status', '')
    platform = _req_platform()
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    accounts, total = models['account'].get_paginated(
        page=page, page_size=page_size,
        keyword=keyword, identity_status=identity_status,
        date_from=date_from, date_to=date_to,
    )

    emails = [acc['email'] for acc in accounts]
    card_counts = models['card_binding'].count_by_emails(emails)
    pa_map = models['platform_account'].map_by_email(platform, emails)
    # 今日/累计成功充值金额，一次聚合出本页全部账号。逐账号查询在 page_size=100 时
    # 会打上百条 SQL——与 card_counts 同一个理由，做法也照抄它。
    recharge_amounts = models['recharge_log'].amount_by_emails(platform, emails)

    # 平台状态过滤只能在这里做：它来自另一张表，塞不进 accounts 的分页 SQL。
    # 代价是分页会与 total 对不齐（本页过滤掉几条，总数仍是身份层的计数）。
    # 这是有意的折中——把平台状态并进分页查询要写成 JOIN + 动态条件，
    # 而这个筛选平时很少用。前端在启用该筛选时提示"总数为身份层计数"。
    if platform_status:
        accounts = [a for a in accounts
                    if (pa_map.get(a['email']) or {}).get('status') == platform_status]

    data = []
    for acc in accounts:
        pa = pa_map.get(acc['email']) or {}
        data.append({
            "email": acc['email'],
            "password": acc.get('login_password') or '',
            # identity_status：GitHub 注册/封禁结果，跨平台一致
            # platform_status：该账号在当前平台的业务状态；'' 表示尚未在此平台开通
            "identity_status": acc.get('identity_status') or '',
            "platform": platform,
            "platform_status": pa.get('status') or '',
            "time": acc.get('created_at') or '',
            "email_password": acc.get('email_password') or '',
            # card_count：库内 card_bindings 成功关联的卡数（可点开看明细）
            "card_count": card_counts.get(acc['email'], 0),
            # 成功充值金额，只算本平台。缺省给 0 而不是 null——列表要直接求和/排序，
            # null 会让前端每个用到的地方都得判空。
            "recharge_today": (recharge_amounts.get(acc['email']) or {}).get('today', 0),
            "recharge_total": (recharge_amounts.get(acc['email']) or {}).get('total', 0),
            "credits_balance": pa.get('credits_balance'),
            "balance_updated_at": pa.get('balance_updated_at') or '',
            "apikey": pa.get('apikey') or '',
            "apikey_updated_at": pa.get('apikey_updated_at') or '',
            "tenant_id": pa.get('tenant_id') or '',
            "email_verify_link": acc.get('email_verify_link') or '',
        })

    return jsonify({"data": data, "total": total, "page": page, "page_size": page_size,
                    "platform": platform})


@api.route('/api/accounts/template')
def download_accounts_template():
    """下载账号导入模版（邮箱 / 邮箱密码 / 邮箱认证链接）。"""
    from src.services import account_import
    path = account_import.generate_template()
    return send_file(path, as_attachment=True, download_name='accounts_template.xlsx')


@api.route('/api/accounts/import', methods=['POST'])
def import_accounts():
    """导入账号 Excel，全部落 identity_status='imported'（待注册）。

    与卡池上传同构（存 uploads/ 再解析），但**不因解析有问题就整批拒绝**：
    能导多少导多少，把问题一起回传。一张几百行的表里有两行格式不对就全批退回，
    只会逼用户来回试。
    """
    from src.services import account_import
    models = get_models()

    if 'file' not in request.files:
        return jsonify({"error": "未上传文件"}), 400
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "仅支持 .xlsx/.xls 文件"}), 400

    upload_dir = str(get_data_dir() / "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    save_path = os.path.join(upload_dir, "accounts_upload.xlsx")
    file.save(save_path)

    rows, errors = account_import.parse_excel(save_path)
    if not rows:
        return jsonify({"error": "没有解析出任何账号", "details": errors}), 400

    stat = account_import.import_rows(models['account'], rows)
    return jsonify({
        "imported": stat['imported'],
        # 没有收码链接的账号入了库却领不走（_hotmail_for_account 会把它们过滤掉），
        # 必须单独回传让前端提示，否则用户会困惑「导入成功了怎么不注册」。
        "no_link": stat['no_link'],
        "no_link_count": len(stat['no_link']),
        "total_parsed": len(rows),
        "errors": errors,
    })


@api.route('/api/accounts/delete', methods=['POST'])
def delete_accounts():
    models = get_models()
    data = request.json or {}
    emails = data.get('emails', [])
    if not emails:
        return jsonify({"error": "没有指定要删除的账号"}), 400

    # 先释放 AdsPower 环境，再删 DB 行。
    #
    # 顺序是刻意的：环境配额只有 12 格，而删账号是「我不要它了」的终局操作——不在这里
    # 释放，那一格就一直被占着，直到下一次撞配额触发 reclaim 才会被当孤儿收掉。
    # 反过来说，AdsPower 挂了/没开也绝不能挡住删账号，所以整段是 best-effort：
    # 任何异常都吞掉，账号照删（残留映射仍会被 reclaim 的第 0 档兜底回收）。
    ads = _release_adspower_for(emails)

    # 删除关联的卡片绑定记录
    placeholders = ','.join(['?'] * len(emails))
    models['card_binding'].db.execute(
        f"DELETE FROM card_bindings WHERE bound_to_email IN ({placeholders})", emails
    )

    # 身份没了，它在各平台的账号行也就没有意义了——不清会留下引用不存在邮箱的孤儿行
    models['platform_account'].delete_all_for_emails(emails)
    count = models['account'].delete_by_emails(emails)
    return jsonify({"deleted": count, "adspower": ads})


@api.route('/api/accounts/archive', methods=['POST'])
def archive_accounts():
    """批量归档：identity_status='retired'，并同步释放 AdsPower 环境。

    `retired` 是身份层终态，表示**用户主动决定不再用这个账号**（区别于
    banned/suspended 那种「账号坏了」）。它在 utils.IDENTITY_TERMINAL_STATUSES 里，
    而四处「还能不能跑」的判据全都只调 is_identity_terminal，所以归档即刻对
    充值/复用/订阅/补号全部生效，不需要逐个平台去标。

    环境一并释放：只有 12 格配额，归档账号还占着就是白占。释放是 best-effort——
    AdsPower 没开/删除失败都不该让归档失败（与 delete_accounts 同一条红线），
    残留映射由 reclaim 的 IDENTITY_DEAD_ORDER 兜底（retired 在其中）。
    """
    models = get_models()
    emails = (request.json or {}).get('emails', [])
    if not emails:
        return jsonify({"error": "没有指定要归档的账号"}), 400

    count = models['account'].set_identity_status(emails, 'retired')
    ads = _release_adspower_for(emails)
    return jsonify({"retired": count, "adspower": ads})


@api.route('/api/accounts/unarchive', methods=['POST'])
def unarchive_accounts():
    """取消归档：retired → registered。

    WHERE 里带 identity_status='retired'，只动归档过的行——批量接口误传几个正常账号
    时不该把它们的状态也改掉。

    环境在归档时已经被删了，所以恢复后首次运行要重新 GitHub 登录，并会触发一次
    新设备邮箱验证。前端提示里要写明这一点。
    """
    models = get_models()
    emails = (request.json or {}).get('emails', [])
    if not emails:
        return jsonify({"error": "没有指定要取消归档的账号"}), 400

    count = models['account'].set_identity_status(
        emails, 'registered', only_from='retired')
    return jsonify({"restored": count})


@api.route('/api/accounts/reset-imported', methods=['POST'])
def reset_accounts_to_imported():
    """批量把注册失败的账号退回 imported，让下一轮重新注册 GitHub。

    判定走 services.account_reset（与命令行脚本共用一份），它挡住两类账号：
      - 状态不是 failed/pending 的（suspended 刻意不可重置，见该模块 docstring）
      - 拿不到收码数据的（重置了也领不走，只会让列表多几行看着能用其实不能用的账号）

    返回三类明细而不是一个数字：用户选了 38 个结果只重置了 12 个时，必须能当场看出
    另外 26 个为什么没动，否则会以为功能坏了。
    """
    from src.services.account_reset import classify_for_reset, load_hotmail_emails

    models = get_models()
    emails = (request.json or {}).get('emails', [])
    if not emails:
        return jsonify({"error": "没有指定要重置的账号"}), 400

    accounts = models['account'].get_by_emails(emails)
    ready, bad_status, no_mailbox = classify_for_reset(
        accounts, load_hotmail_emails())

    if ready:
        models['account'].set_identity_status(
            [a['email'] for a in ready], 'imported')
    return jsonify({
        "reset": [a['email'] for a in ready],
        "skipped_status": [
            {"email": a['email'], "status": a.get('identity_status') or ''}
            for a in bad_status],
        "skipped_no_mailbox": [a['email'] for a in no_mailbox],
    })


def _release_adspower_for(emails):
    """删账号时同步删掉它们的 AdsPower 环境，返回给前端看的结果摘要。

    绝不抛异常：调用方是删除/归档接口，环境释放失败不该让账号删不掉、归不了档。
    """
    result = {"released": [], "skipped_busy": [], "failed": [], "reason": ""}
    state = get_app_state()
    try:
        if not state.adspower_enabled:
            result["reason"] = "AdsPower 未启用，跳过环境释放"
            return result
        # 复用共享的池而不是新建：它的 _lock 串行化着「挑代理→建环境→撞配额→回收」，
        # 新建一个等于绕开那把锁，可能与正在建环境的 worker 撞在一起。
        _client, pool = state._ensure_adspower()
        if pool is None:
            result["reason"] = "AdsPower 环境池不可用，跳过环境释放"
            return result
        out = pool.release_many(emails)
        result.update({k: out.get(k, []) for k in
                       ("released", "skipped_busy", "failed")})
        if result["skipped_busy"]:
            result["reason"] = "部分账号正在运行，其环境保留至跑完后自动回收"
        elif result["failed"]:
            result["reason"] = "AdsPower 删除环境失败，环境仍占用配额"
    except Exception as e:      # noqa: BLE001 —— 见 docstring：绝不阻断删除
        result["failed"] = list(emails)
        result["reason"] = f"释放 AdsPower 环境时出错: {str(e)[:150]}"
        state.add_log(f"[AdsPower] 删账号时释放环境失败: {str(e)[:150]}")
    return result


@api.route('/api/accounts/<email>/cards')
def get_account_cards(email):
    models = get_models()
    cards = models['card_binding'].get_by_email(email)
    return jsonify(cards)


def _recharge_cfg_from(data):
    """从请求体解析充值策略覆盖，返回 (RechargeConfig, err)。err 非空时应回 400。

    校验放在这一层做一次、模型层不重复。**非法值返回 400 而不是静默夹紧**：
    用户配的是 20-100、实际跑的却是别的，比直接报错难查得多。
    只有越界（超出 RechargeConfig 的绝对边界）才夹紧——那是防呆，不是改语义。

    三个字段都可缺省；全缺省时返回 cfg.recharge 本身（不复制，因为没改动）。
    """
    keys = ('amount_min', 'amount_max', 'balance_cap')
    if not any(data.get(k) is not None for k in keys):
        return cfg.recharge, ''

    try:
        amount_min = None if data.get('amount_min') is None else int(data['amount_min'])
        amount_max = None if data.get('amount_max') is None else int(data['amount_max'])
        balance_cap = None if data.get('balance_cap') is None else float(data['balance_cap'])
    except (TypeError, ValueError):
        return None, "充值金额与余额上限必须是数字"

    # 只给一端时另一端沿用当前配置，否则「只改下界」会变成一个倒挂区间。
    lo = amount_min if amount_min is not None else cfg.recharge.amount_min
    hi = amount_max if amount_max is not None else cfg.recharge.amount_max
    if lo > hi:
        return None, f"充值金额区间非法：下界 ${lo} 大于上界 ${hi}"
    if lo < RechargeConfig.AMOUNT_FLOOR or hi > RechargeConfig.AMOUNT_CEILING:
        return None, (f"充值金额需在 ${RechargeConfig.AMOUNT_FLOOR}–"
                      f"${RechargeConfig.AMOUNT_CEILING} 之间")
    if balance_cap is not None and balance_cap <= 0:
        return None, "账号余额上限必须大于 0"

    return cfg.recharge.with_overrides(
        amount_min=lo, amount_max=hi, balance_cap=balance_cap), ''


@api.route('/api/accounts/recharge', methods=['POST'])
def recharge_account():
    models = get_models()

    data = request.json or {}
    platform = _req_platform()
    # 按平台取运行上下文——闸门、计数、日志全都是这个 ctx 的。
    # 曾经这里取的是全局单例，于是 infron 的充值会被 opencode 正在跑的任务挡住。
    state = get_ctx(platform)

    if state.is_running:
        return jsonify({"error": f"{platform} 已有任务在运行，请等待完成后再操作"}), 400
    email = data.get('email', '')
    payment_group_id = data.get('payment_group_id')
    captcha_api_key = data.get('captcha_api_key')
    # 充值默认用 Multibot 解支付页 hCaptcha；可传 captcha_server='2captcha.com' 切回
    captcha_server = data.get('captcha_server') or 'api.multibot.cloud'
    recharge_cfg, cfg_err = _recharge_cfg_from(data)
    if cfg_err:
        return jsonify({"error": cfg_err}), 400
    if not email:
        return jsonify({"error": "未指定账号"}), 400

    # 查找账号
    rows = models['account'].search(email)
    account = None
    for r in rows:
        if r['email'] == email:
            account = r
            break

    if not account:
        return jsonify({"error": "账号不存在"}), 404

    # 不在这里校验凭据：**需要哪些凭据是平台决定的**。opencode 走 GitHub OAuth 要
    # login_password，infron 走邮箱 magic link 只要 email_verify_link，将来的平台可能
    # 两个都不要。缺什么由 adapter.ensure_session 判断并给出可读的 detail
    # （opencode 会说「该账号未保存登录密码，且 profile 未登录」）。
    #
    # 这个前置校验对 opencode 也偏严：profile 已登录但没存密码的账号本来能跑，
    # 却被挡在门外。
    login_password = account.get('login_password')

    # opencode 充值时直接在 Stripe Checkout 填卡，无需预先绑卡，故不再校验 bound 状态。
    # 支付卡来自 payment_group_id 指定的卡池分组（见 _recharge_one_account → recharge_account）。

    # 不再需要「把 platform 赋给 state」——ctx 本身就是按平台取的，它的 .platform
    # 就是这个值。曾经那行赋值是单例时代的产物，多平台并发下改全局字段正是要消除的
    # 竞态源（两个平台同时跑会互相覆盖）。

    # 标记为运行中，在后台线程执行充值。
    #
    # stop_requested 必须跟着一起复位：它是**跨任务残留**的。上一次任务被用户停止、
    # 或某个 worker 抛 InterruptedError（worker.py 会置全局 stop_requested 让其它
    # worker 一起收敛）之后，这个标志一直是 True；不复位的话，下一次充值会在第一个
    # 检查点就自杀，日志只留一句「收到停止请求，正在中断」——看起来像用户又点了停止，
    # 极难往「上一轮的残留」上想。三条流水线入口都成对复位了，只有这里漏了。
    state.is_running = True
    state.stop_requested = False
    state.current_action = f"正在为 {email} 在 {platform} 充值..."

    import threading

    def _do_recharge():
        # 绑定必须在**本线程内**做：contextvars 不跨线程继承，在请求线程里绑
        # 对这个新线程毫无作用，日志会因为解析不出归属而退化成裸 print。
        # （_patch_prints 同时完成「装钩子」与「绑本线程」，钩子部分是幂等的。）
        state._patch_prints()
        try:
            state._recharge_one_account(email, login_password, payment_group_id,
                                        captcha_api_key=captcha_api_key,
                                        captcha_server=captcha_server,
                                        recharge_cfg=recharge_cfg)
        except InterruptedError:
            state.add_log("充值已中断")
        except Exception as e:
            state.add_log(f"充值异常: {e}")
        finally:
            state.clear_active_driver()
            state._stop_screenshot_loop()
            state.is_running = False

    threading.Thread(target=_do_recharge, daemon=True).start()
    return jsonify({"status": "started", "email": email})


@api.route('/api/accounts/open-browser', methods=['POST'])
def open_account_browser():
    """打开浏览器查看账号（平台控制台），按账号独立，不阻塞全局任务"""
    models = get_models()
    platform = _req_platform()
    # 按平台取：下游要用这个 ctx 的 browser_factory 与该平台的适配器。
    state = get_ctx(platform)

    data = request.json or {}
    email = data.get('email', '')
    if not email:
        return jsonify({"error": "未指定账号"}), 400

    rows = models['account'].search(email)
    account = None
    for r in rows:
        if r['email'] == email:
            account = r
            break

    if not account:
        return jsonify({"error": "账号不存在"}), 404

    # 不在这里校验凭据：**需要哪些凭据是平台决定的**。opencode 走 GitHub OAuth 要
    # login_password，infron 走邮箱 magic link 只要 email_verify_link，将来的平台可能
    # 两个都不要。缺什么由 adapter.ensure_session 判断并给出可读的 detail
    # （opencode 会说「该账号未保存登录密码，且 profile 未登录」）。
    #
    # 这个前置校验对 opencode 也偏严：profile 已登录但没存密码的账号本来能跑，
    # 却被挡在门外。
    login_password = account.get('login_password')

    # 预约 profile：与任务 worker 的账号占用共用一把锁，避免同一 Chrome profile
    # 被 worker 与手动会话同时使用（会互删 Singleton 锁导致浏览器崩溃）
    ok, reason = state.account_registry.try_open_manual(email)
    if not ok:
        return jsonify({"error": reason}), 409

    state.add_log(f"{email} 正在打开浏览器...")

    import threading

    def _do_open():
        from src.browser.driver import create_driver, close_driver
        from src.services.adspower import AdsPowerError
        import src.platforms as platforms
        from src.platforms.base import Credentials
        adapter = platforms.get(platform)
        driver = None
        try:
            # 开哪个浏览器必须与**任务用的那个**一致，否则「查看」看到的根本不是
            # 任务跑的那个账号环境：
            #   AdsPower 开启时，登录态（GitHub cookie / opencode session）全在该账号
            #   的 AdsPower 环境里；本地 data/profiles/<email> 是另一个几乎空的目录。
            #   拿本地 profile 打开，用户看到的是未登录页，ensure_session 还会在这个
            #   错误的环境里重新走一遍 OAuth——既看不到真实状态，又平白给账号多一次
            #   新设备登录记录。
            # 走同一个 pool 就能拿到同一个 profile_id（pool 按 email 映射），环境复用、
            # 登录态自然还在。
            factory = state.browser_factory(track_for_teardown=False)
            if factory is not None:
                driver = factory(email)
                state.add_log(f"{email} 已接管其 AdsPower 环境")
            else:
                driver = create_driver(headless=False, profile_id=email)

            # 复用充值流程用的同一套会话建立逻辑（adapter.ensure_session）。
            # 遇新设备验证等人工环节，该流程会保持浏览器打开等待人工完成。
            def _monitor(_drv, step):
                if step:
                    state.add_log(f"{email}: {step}")

            sess = adapter.ensure_session(
                driver,
                Credentials(email=email, login_password=login_password,
                            verify_link=account.get('email_verify_link')),
                monitor=_monitor,
            )
            if sess.ok:
                state.add_log(f"{email} 已登录 {adapter.display_name}"
                              f"（{sess.detail}），租户={sess.tenant_id}")
            else:
                state.add_log(f"{email} 未能自动登录 {adapter.display_name}："
                              f"{sess.detail}，请在浏览器手动操作")

            # 等待用户手动关闭浏览器；期间轮询 billing 页 Current Balance，
            # 用户进入结算页时读到即落库刷新余额（driver.title 触发事件派发保活）。
            import time
            last_persisted = None
            while True:
                try:
                    _ = driver.title
                    bal = adapter.read_balance_from_current_page(driver)
                    if bal is not None and bal != last_persisted:
                        try:
                            models['platform_account'].update_balance(platform, email, bal)
                            last_persisted = bal
                            state.add_log(f"{email} 检测到余额页，余额已更新: ${bal:.2f}")
                        except Exception as _e:
                            state.add_log(f"{email} 余额落库失败: {str(_e)[:80]}")
                    time.sleep(2)
                except Exception:
                    break
        except AdsPowerError as e:
            # 环境配额满 / 客户端没开 / Key 无效。单独报是因为这几种情况用户自己
            # 就能处理（关掉别的环境、开客户端），而混进通用异常里只会看到一串栈。
            state.add_log(f"{email} 打不开 AdsPower 环境: {e}")
        except Exception as e:
            state.add_log(f"{email} 打开浏览器异常: {e}")
        finally:
            if driver:
                try:
                    close_driver(driver)
                except Exception:
                    pass
            state.open_browsers.discard(email)
            state.add_log(f"{email} 浏览器已关闭")

    threading.Thread(target=_do_open, daemon=True).start()
    return jsonify({"status": "started", "email": email})


@api.route('/api/accounts/open-browsers')
def get_open_browsers():
    """获取当前打开的浏览器会话列表。

    open_browsers 是**共享**资源（Chrome profile 目录按 email，跨平台唯一），
    取哪个 ctx 拿到的都是同一个集合——这里用默认 ctx 不是漏改。
    """
    state = get_app_state()
    return jsonify({"emails": list(state.open_browsers)})


@api.route('/api/recharge-logs')
def get_recharge_logs():
    models = get_models()
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 20))
    email = request.args.get('email', '')
    status = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    logs, total = models['recharge_log'].get_paginated(
        page=page, page_size=page_size,
        email=email, status=status,
        date_from=date_from, date_to=date_to,
    )
    return jsonify({"data": logs, "total": total, "page": page, "page_size": page_size})


@api.route('/api/card-recharge-logs')
def get_recharge_logs_by_card():
    """某张卡的充值记录明细，供卡池/有效卡列表点开查看。

    路径特意不挂在 /api/recharge-logs/ 下——那里的 <email> 动态段会吞掉同级路径。
    """
    models = get_models()
    card_number = request.args.get('card_number', '')
    if not card_number:
        return jsonify({"error": "缺少 card_number"}), 400
    logs = models['recharge_log'].get_by_card(card_number)
    success = sum(1 for l in logs if l.get('status') == 'success')
    return jsonify({"data": logs, "total": len(logs), "success_total": success,
                    "card_number": card_number})


@api.route('/api/recharge-logs/<email>')
def get_recharge_logs_by_email(email):
    models = get_models()
    logs = models['recharge_log'].get_by_email(email)
    return jsonify(logs)


# 报表区间的默认跨度（含今天），与前端「重置」回落到的区间保持一致
REPORT_DEFAULT_DAYS = 30
# 账号榜单展示上限。只截断**展示**，汇总与核销拆分一律在全量上算
REPORT_ACCOUNT_LIMIT = 100


def _report_range(args):
    """取报表区间 (date_from, date_to)，均为 'YYYY-MM-DD'。

    任一端缺省就补成「最近 30 天」——不是「全时段」。全时段作为默认会让首次打开
    报表页就扫全表，且日趋势表拉出几百行没人看；用户要看更早的数据，把日期往前拨即可。

    日期用 datetime.date.today()：它取的是本机本地日期，与 recharge_logs 里
    datetime('now','localtime') 写入的时间戳同一时区，不会差出一天。
    """
    from datetime import date, timedelta
    date_from = (args.get('date_from') or '').strip()
    date_to = (args.get('date_to') or '').strip()
    if not date_to:
        date_to = date.today().isoformat()
    if not date_from:
        date_from = (date.today() - timedelta(days=REPORT_DEFAULT_DAYS - 1)).isoformat()
    return date_from, date_to


@api.route('/api/reports/recharge')
def get_recharge_report():
    """充值报表：今日 KPI + 区间汇总 + 逐日趋势 + 账号榜单。

    口径见 RechargeLogModel 报表区的注释——只算 status='success'，一律按 platform 过滤。
    `today` 段不受区间参数影响：KPI 卡显示的是「今天」，用户把区间拉到上个月时它不该归零。
    """
    models = get_models()
    platform = _req_platform()
    date_from, date_to = _report_range(request.args)

    rlog = models['recharge_log']
    summary = rlog.report_summary(platform, date_from, date_to)
    today = rlog.report_today(platform)
    daily = rlog.report_daily(platform, date_from, date_to)
    # 取全量：verified/active 的拆分必须在全量上做，否则两段金额之和小于 summary.total_amount
    accounts = rlog.report_by_account(platform, date_from, date_to)

    # 成功率在后端算：前端有三处要显示它（KPI、汇总条、逐日表），
    # 分母为 0 的处理散在三个模板里迟早会漏一处除零。
    attempts = summary['success_count'] + summary['failed_count']
    summary['success_rate'] = round(summary['success_count'] / attempts * 100, 1) if attempts else 0.0

    # 「已核销」判据来自 accounts 表的身份状态，跨表拼在 Python 里而不是 SQL JOIN：
    # 榜单最多 100 行，一次 get_by_emails 的成本可忽略，而 JOIN 会把「账号身份」
    # 这个概念泄进充值模型，那里本不该知道 retired 是什么。
    status_map = {
        a['email']: (a.get('identity_status') or '')
        for a in models['account'].get_by_emails([r['email'] for r in accounts])
    }
    verified = {'amount': 0.0, 'account_count': 0}
    active = {'amount': 0.0, 'account_count': 0}
    for row in accounts:
        row['identity_status'] = status_map.get(row['email'], '')
        # 前端只读 is_verified，不自己比对字符串——核销判据将来若扩到多个状态，只改这一处
        row['is_verified'] = row['identity_status'] == 'retired'
        bucket = verified if row['is_verified'] else active
        bucket['amount'] += row['amount']
        bucket['account_count'] += 1
    verified['amount'] = round(verified['amount'], 2)
    active['amount'] = round(active['amount'], 2)

    return jsonify({
        "platform": platform,
        "date_from": date_from,
        "date_to": date_to,
        "today": today,
        "summary": summary,
        "verified": verified,
        "active": active,
        "daily": daily,
        # 榜单只展示前 100 名；上面的拆分与汇总都已在全量上算完，截断不影响任何数字
        "accounts": accounts[:REPORT_ACCOUNT_LIMIT],
        "accounts_total": len(accounts),
    })


@api.route('/api/accounts/export', methods=['POST'])
def export_accounts():
    models = get_models()
    data = request.json or {}
    platform = _req_platform()

    mode = data.get('mode', 'filtered')
    selected_emails = data.get('emails', [])

    if mode == 'selected' and selected_emails:
        accounts = []
        for email in selected_emails:
            rows = models['account'].search(email)
            for r in rows:
                if r['email'] == email:
                    accounts.append(r)
                    break
    else:
        keyword = data.get('keyword', '')
        identity_status = data.get('identity_status', '')
        date_from = data.get('date_from', '')
        date_to = data.get('date_to', '')
        accounts, _ = models['account'].get_paginated(
            page=1, page_size=99999,
            keyword=keyword, identity_status=identity_status,
            date_from=date_from, date_to=date_to,
        )

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"

    # apikey 与余额是平台数据，导出的是当前选定平台那一份
    pa_map = models['platform_account'].map_by_email(
        platform, [a['email'] for a in accounts])

    headers = ["邮箱", "GitHub密码", "邮箱密码", "认证链接",
               f"apikey({platform})", f"余额({platform})"]
    header_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    header_font = Font(bold=True, size=11)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    row_idx = 2
    for acc in accounts:
        pa = pa_map.get(acc['email']) or {}
        ws.cell(row=row_idx, column=1, value=acc['email'])
        ws.cell(row=row_idx, column=2, value=acc.get('login_password') or '')
        ws.cell(row=row_idx, column=3, value=acc.get('email_password') or '')
        ws.cell(row=row_idx, column=4, value=acc.get('email_verify_link') or '')
        ws.cell(row=row_idx, column=5, value=pa.get('apikey') or '')
        ws.cell(row=row_idx, column=6, value=pa.get('credits_balance'))
        row_idx += 1

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='accounts_export.xlsx',
    )


# ==================== 卡片分组 & 底料池 ====================

@api.route('/api/card-groups')
def get_card_groups():
    models = get_models()
    group_type = request.args.get('type', '')
    if group_type:
        groups = models['card_group'].get_by_type(group_type)
    else:
        groups = models['card_group'].get_all()
    return jsonify(groups)


@api.route('/api/card-groups', methods=['POST'])
def create_card_group():
    models = get_models()
    data = request.json or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({"error": "分组名称不能为空"}), 400
    # 已不区分分组类型，统一支付卡；type 仅为兼容保留，默认 payment
    group_type = data.get('type') or 'payment'
    if group_type not in ('bind', 'payment'):
        group_type = 'payment'
    description = data.get('description', '')
    group_id = models['card_group'].create(name, group_type, description)
    return jsonify({"id": group_id, "name": name, "type": group_type})


@api.route('/api/card-groups/<int:group_id>', methods=['PUT'])
def update_card_group(group_id):
    models = get_models()
    data = request.json or {}
    name = data.get('name')
    description = data.get('description')
    models['card_group'].update(group_id, name=name, description=description)
    return jsonify({"status": "ok"})


@api.route('/api/card-groups/<int:group_id>', methods=['DELETE'])
def delete_card_group(group_id):
    models = get_models()
    models['card_group'].delete(group_id)
    return jsonify({"status": "ok"})


def _attach_recharge_counts(card, counts):
    """给一张卡补充 recharge_total（累计成功充值次数）/ recharge_today（当日成功充值次数）。
    counts 由 RechargeLogModel.count_success_by_last4() 一次性聚合得到，按卡号末 4 位匹配。"""
    last4 = (card.get('card_number') or '').replace(' ', '')[-4:]
    c = counts.get(last4) or {}
    card['recharge_total'] = c.get('total', 0)
    card['recharge_today'] = c.get('today', 0)
    return card


@api.route('/api/card-pool/<int:group_id>')
def get_card_pool(group_id):
    models = get_models()
    platform = _req_platform()
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 20))
    bucket = request.args.get('bucket', '')  # ''=全部 / valid / unverified / invalid

    # 先按当前日期刷新过期状态，列表里直接能看到哪些卡已过期
    models['card_pool'].refresh_expired_status(group_id)
    cards, total = models['card_pool'].get_by_group(
        platform, group_id, page=page, page_size=page_size, bucket=bucket)

    # 标记有效卡 + 选卡规则状态（供列表状态列展示冷却与连续失败次数）。
    # 全部按当前平台算——同一张卡在别的平台的冷却与绑定跟这个列表无关。
    #
    # 这里**不再**报「24h 内成功 ≥2 次」那个标记。它从来没有真正参与选卡
    # （_eligible_cards 只看 card_state 的冷却），而现在同一张卡在一个账号的会话里
    # 连着成功多笔本就是预期行为，把它显示成冷却会让整列好卡看起来都出了问题。
    # 取而代之的是连续失败次数——那才是「这张卡快被判废了」的真信号。
    state_map = models['card_state'].get_state_map(platform)
    recharge_counts = models['recharge_log'].count_success_by_last4(platform)
    for card in cards:
        num = card['card_number']
        st = state_map.get(num) or {}
        card['is_valid'] = models['valid_card'].is_valid(platform, num)
        card['tds_cooldown'] = bool(st.get('in_cooldown'))
        card['tds_until'] = st.get('tds_until')
        card['fail_streak'] = st.get('fail_streak') or 0
        card['max_fail_streak'] = cfg.recharge.fail_threshold()
        card['bound_email'] = models['valid_card'].get_bound_email(platform, num)
        _attach_recharge_counts(card, recharge_counts)

    buckets = models['card_pool'].count_buckets(platform, group_id)
    return jsonify({"data": cards, "total": total, "page": page, "page_size": page_size,
                    "bucket": bucket, "buckets": buckets, "platform": platform})


@api.route('/api/card-pool/merge', methods=['POST'])
def merge_card_pools():
    """把多个源分组里的"非无效"卡（有效+未验证）移动合并到一个新分组。"""
    models = get_models()
    data = request.json or {}
    source_ids = data.get('source_group_ids') or []
    name = (data.get('name') or '').strip()
    # 已不区分分组类型，统一支付卡
    group_type = data.get('type') or 'payment'
    if group_type not in ('bind', 'payment'):
        group_type = 'payment'
    # 转移范围：non_invalid（有效+未验证，默认）/ valid（仅有效卡）
    bucket = data.get('bucket') or 'non_invalid'
    if bucket not in ('non_invalid', 'valid'):
        return jsonify({"error": "转移范围无效"}), 400
    if not source_ids:
        return jsonify({"error": "未选择源分组"}), 400
    if not name:
        return jsonify({"error": "未填写新分组名称"}), 400
    # 源分组存在性校验
    for gid in source_ids:
        if not models['card_group'].get_by_id(gid):
            return jsonify({"error": f"源分组 {gid} 不存在"}), 404

    new_id = models['card_group'].create(name, group_type=group_type)
    result = models['card_pool'].move_non_invalid_to_group(
        _req_platform(), source_ids, new_id, bucket=bucket)
    return jsonify({"status": "ok", "group_id": new_id,
                    "moved": result['moved'], "deduped": result['deduped']})


@api.route('/api/card-pool/<int:group_id>/move', methods=['POST'])
def move_cards_to_group(group_id):
    """把源分组内指定桶的卡片，按数量上限移动到已存在的目标分组。

    与 /merge 的区别：目标分组必须已存在，可限制数量与桶，重复卡跳过而非删除。
    """
    models = get_models()
    data = request.json or {}

    target_id = data.get('target_group_id')
    if not isinstance(target_id, int) or isinstance(target_id, bool):
        return jsonify({"error": "未选择目标分组"}), 400
    if target_id == group_id:
        return jsonify({"error": "源分组与目标分组相同"}), 400

    bucket = data.get('bucket')
    if bucket not in CardPoolModel.MOVABLE_BUCKETS:
        return jsonify({"error": "卡片范围无效"}), 400

    limit = data.get('limit')
    if not isinstance(limit, int) or isinstance(limit, bool) or limit <= 0:
        return jsonify({"error": "移动数量必须为正整数"}), 400

    if not models['card_group'].get_by_id(group_id):
        return jsonify({"error": "源分组不存在"}), 404
    if not models['card_group'].get_by_id(target_id):
        return jsonify({"error": "目标分组不存在"}), 404

    result = models['card_pool'].move_bucket_to_group(
        _req_platform(), group_id, target_id, bucket, limit)
    return jsonify({"status": "ok", "moved": result['moved'], "skipped": result['skipped']})


@api.route('/api/card-pool/<int:group_id>/delete-invalid', methods=['POST'])
def delete_invalid_cards(group_id):
    """删除某分组内所有无效卡（invalid + expired）。"""
    models = get_models()
    if not models['card_group'].get_by_id(group_id):
        return jsonify({"error": "分组不存在"}), 404
    deleted = models['card_pool'].delete_invalid_by_group(_req_platform(), group_id)
    return jsonify({"status": "ok", "deleted": deleted})


@api.route('/api/card-pool/<int:group_id>/upload', methods=['POST'])
def upload_card_pool(group_id):
    models = get_models()

    group = models['card_group'].get_by_id(group_id)
    if not group:
        return jsonify({"error": "分组不存在"}), 404

    if 'file' not in request.files:
        return jsonify({"error": "未上传文件"}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "仅支持 .xlsx/.xls 文件"}), 400

    upload_dir = str(get_data_dir() / "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    save_path = os.path.join(upload_dir, f"pool_upload_{group_id}.xlsx")
    file.save(save_path)

    cards, errors = card_service.parse_excel(save_path)
    if not cards:
        return jsonify({"error": "解析失败", "details": errors}), 400

    added, skipped, conflicts = models['card_pool'].add_cards(group_id, cards)

    # 上传时即按有效期判定，过期卡会入库但标记为 expired（不参与任务），这里回传数量供前端提示
    expired = sum(1 for c in cards if is_card_expired(c.get('expiry_month'), c.get('expiry_year')))

    return jsonify({
        "added": added,
        "skipped": skipped,
        "conflicts": conflicts,
        "expired": expired,
        "total_parsed": len(cards),
        "errors": errors,
    })


@api.route('/api/card-pool/card/<int:card_id>', methods=['DELETE'])
def delete_pool_card(card_id):
    models = get_models()
    models['card_pool'].delete_card(card_id)
    return jsonify({"status": "ok"})


@api.route('/api/card-pool/<int:group_id>/clear', methods=['POST'])
def clear_card_pool(group_id):
    models = get_models()
    deleted = models['card_pool'].delete_by_group(group_id)
    return jsonify({"deleted": deleted})


# ==================== 有效卡 ====================

_POOL_STATUS_ZH = {'': '在库(未验证)', 'paid': '有效(已支付)', 'invalid': '无效',
                   'expired': '已过期', 'bound': '已绑定'}


def _valid_card_status(models, card, platform, state_map=None):
    """给一张有效卡补充选卡状态：绑定账号、冷却、连续失败次数、汇总状态文案、池内分组/状态。

    全部按 platform 算：冷却、失败计数、池内状态在各平台各不相同。

    与列表接口同口径：不再报「24h 内成功 ≥2 次」——它不参与选卡，而同一张卡连着
    成功多笔现在是预期行为，标成冷却只会误导。

    state_map: `card_state.get_state_map(platform)` 的结果。**调用方必须在循环外取一次
    并传进来**——它是该平台的整表扫描，放在这里现取的话，导出接口（全量有效卡、可上千行）
    会把它跑成 O(卡数²)。省略时退化为自取，只为单卡调用方便。"""
    num = card.get('card_number', '')
    if state_map is None:
        state_map = models['card_state'].get_state_map(platform)
    st = state_map.get(num) or {}
    tds = bool(st.get('in_cooldown'))
    streak = st.get('fail_streak') or 0
    cap = cfg.recharge.fail_threshold()
    card['bound_email'] = card.get('source_email', '') if card.get('source_type') == 'payment' else ''
    card['tds_cooldown'] = tds
    card['tds_until'] = st.get('tds_until')
    card['fail_streak'] = streak
    card['max_fail_streak'] = cap
    card['status_text'] = ('冷却中' if tds
                           else (f'连续失败 {streak}/{cap}' if streak else '可用'))
    # 池内位置：该有效卡当前在卡池哪个分组、什么状态（解释"为何不计入某分组的有效桶"）
    locs = models['card_pool'].get_locations_by_number(platform, num)
    if locs:
        card['pool_group'] = '，'.join((l.get('group_name') or str(l.get('group_id'))) for l in locs)
        card['pool_status'] = '，'.join(_POOL_STATUS_ZH.get(l.get('status') or '', l.get('status') or '') for l in locs)
    else:
        card['pool_group'] = ''
        card['pool_status'] = '不在卡池'
    return card


@api.route('/api/valid-cards')
def get_valid_cards():
    models = get_models()
    platform = _req_platform()
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 20))
    source_type = request.args.get('source_type', '')
    keyword = request.args.get('keyword', '')
    cards, total = models['valid_card'].get_all(
        platform, page=page, page_size=page_size,
        source_type=source_type, keyword=keyword,
    )
    recharge_counts = models['recharge_log'].count_success_by_last4(platform)
    state_map = models['card_state'].get_state_map(platform)
    for c in cards:
        _valid_card_status(models, c, platform, state_map)
        _attach_recharge_counts(c, recharge_counts)
    summary = models['valid_card'].get_summary(platform)
    return jsonify({"data": cards, "total": total, "page": page, "page_size": page_size,
                    "summary": summary, "platform": platform})


@api.route('/api/valid-cards/export')
def export_valid_cards():
    """导出全部有效卡为 xlsx：中文表头 + 关联账号信息 + 完整信用卡信息（不脱敏）。"""
    import openpyxl
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    models = get_models()
    platform = _req_platform()
    source_type = request.args.get('source_type', '')
    cards = models['valid_card'].get_all_for_export(platform, source_type)

    # 关联账号信息：身份层取密码，平台层取该平台的状态
    acct_map = {a['email']: a for a in models['account'].get_all(order_desc=False)}
    pa_map = models['platform_account'].map_by_email(platform)

    headers = ['卡号', '有效期(月)', '有效期(年)', '安全码CVC', '名', '姓',
               '国家', '地址', '地址2', '城市', '州', '邮编', '公司',
               '来源', '关联账号', 'GitHub密码', '邮箱密码', '身份状态',
               f'平台状态({platform})',
               '卡状态', '累计充值次数', '当日充值次数', '验证时间']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '有效卡'
    ws.append(headers)
    recharge_counts = models['recharge_log'].count_success_by_last4(platform)
    state_map = models['card_state'].get_state_map(platform)
    for c in cards:
        _valid_card_status(models, c, platform, state_map)
        _attach_recharge_counts(c, recharge_counts)
        email = c.get('source_email', '') or ''
        acct = acct_map.get(email, {})
        ws.append([
            c.get('card_number', ''), c.get('expiry_month', ''), c.get('expiry_year', ''),
            c.get('cvc', ''), c.get('first_name', ''), c.get('last_name', ''),
            c.get('country', ''), c.get('address', ''), c.get('address2', ''),
            c.get('city', ''), c.get('state', ''), c.get('zip', ''), c.get('company', ''),
            ('绑定' if c.get('source_type') == 'bind' else '支付'),
            email, acct.get('login_password', ''), acct.get('email_password', ''),
            acct.get('identity_status', ''),
            (pa_map.get(email) or {}).get('status', ''),
            c.get('status_text', ''),
            c.get('recharge_total', 0), c.get('recharge_today', 0), c.get('validated_at', ''),
        ])

    # 列宽自适应（按每列最长内容估算，含表头）
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        maxlen = 0
        for cell in ws[letter]:
            v = '' if cell.value is None else str(cell.value)
            # 中文按 2 宽度估算
            w = sum(2 if ord(ch) > 127 else 1 for ch in v)
            maxlen = max(maxlen, w)
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 8), 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"有效卡_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@api.route('/api/daily/start', methods=['POST'])
def start_daily_pipeline():
    """启动每日充值任务：选定卡池分组，逐账号轮转充值直到卡池消耗完。"""
    models = get_models()
    data = request.json or {}

    group_id = data.get('group_id')
    if not group_id:
        return jsonify({"error": "未指定卡池分组"}), 400
    group = models['card_group'].get_by_id(group_id)
    if not group:
        return jsonify({"error": "卡池分组不存在"}), 404

    platform = _req_platform(required=True)
    if not platform:
        return jsonify({"error": "未指定平台"}), 400

    # 闸门必须**在拿到 platform 之后**判，而且只判这个平台。
    # 曾经是先取全局单例再判 is_running，于是一个平台在跑就挡住所有平台——
    # 这正是 AC1 要解开的那道锁。
    state = get_ctx(platform)
    if state.is_running:
        return jsonify({"error": f"{platform} 已有任务在运行"}), 400

    login_password = data.get('login_password') or None
    captcha_api_key = data.get('captcha_api_key')
    # 充值默认用 Multibot 解支付页 hCaptcha；可传 captcha_server='2captcha.com' 切回
    captcha_server = data.get('captcha_server') or 'api.multibot.cloud'

    recharge_cfg, cfg_err = _recharge_cfg_from(data)
    if cfg_err:
        return jsonify({"error": cfg_err}), 400

    # 启动门：分组要有可选卡（排除无效/过期/冷却），且要有可充值账号。
    # 账号判据必须与 run_daily_pipeline._payable_now 用同一组常量——此前这里只排除
    # ('banned','archived') 而流水线排除四项，于是「启动时说有 N 个可充值账号，
    # 实际跑起来只有 M 个」。
    # platform 必须显式传下去：_eligible_cards 在 platform=None 时回落 AppState.platform，
    # 也就是**上一次运行**的那个平台。单平台下「碰巧对」，于是这个启动门一直在用
    # 另一个平台的卡数做判断。exclude_used=False 同样是必须的——计数调用点不能扣掉
    # 「本次运行已用」，否则报给用户的可选卡数会偏小。
    eligible = len(state._eligible_cards(group_id, exclude_used=False, platform=platform))
    if not eligible:
        return jsonify({"error": "该分组无可选卡（全部无效/过期或冷却中），无事可做"}), 400

    accts = models['account'].get_all(order_desc=False)
    pa_map = models['platform_account'].map_by_email(platform)

    def _usable(a):
        return ((login_password or a.get('login_password'))
                and not is_identity_terminal(a.get('identity_status')))

    account_count = sum(
        1 for a in accts
        if _usable(a)
        and not is_platform_terminal((pa_map.get(a['email']) or {}).get('status'))
    )
    # 回退池：余额未满的已充值账号。必须算进启动门——它们平台状态是 recharged
    # （终态），上面那个 count 看不见它们，而「新号跑完只剩老号加码」正是复用功能
    # 存在的场景。漏了这一项，接口会在那个场景下直接 400，流水线根本进不去。
    # 判据与 run_daily_pipeline._reusable_recharged 保持一致，改一处要改两处。
    reusable_count = sum(
        1 for a in accts
        if _usable(a)
        and ((pa_map.get(a['email']) or {}).get('status') or '') == 'recharged'
        and ((pa_map.get(a['email']) or {}).get('credits_balance') is None
             or (pa_map.get(a['email']) or {}).get('credits_balance') < recharge_cfg.balance_cap)
    )
    # 待注册 imported 账号：刚导入、GitHub 还没注册的邮箱。**不能用 _usable 判**——
    # 它要求 login_password 非空，而那个密码正是注册流程写回去的，imported 账号
    # 天然没有。用 _usable 判的结果就是「账号列表里只有新导入的邮箱」这个最常见的
    # 开局场景永远起不来，而流水线其实完全跑得动：_try_claim 领不到可充账号时会领
    # 一个 imported 走 _register_one_account，注册成功者下一轮即以 registered 身份充值。
    #
    # 判据与 run_daily_pipeline._registerable_imported() 保持一致，改一处要改两处。
    # 那边多一个 `not in done`（本次运行的终结集合），启动门在运行前判、done 恒空，故略去。
    # 收码数据这一条不能省：没有它的 imported 账号流水线根本领不走，光放行会把
    # 「启动就被拒」变成「启动成功但一轮空跑就收敛」，比 400 更难查。
    registerable_count = sum(
        1 for a in accts
        if (a.get('identity_status') or '') == 'imported'
        and state._hotmail_for_account(a)
    )
    if account_count == 0 and registerable_count == 0 and reusable_count == 0:
        return jsonify({"error": "无可充值账号（需有登录密码、身份与平台状态均非终态）、"
                                 "无待注册 imported 账号（需有收码链接）、"
                                 "也无余额未满的已充值账号可复用，无事可做"}), 400

    import threading
    threading.Thread(
        target=state.run_daily_pipeline,
        # ⚠️ 位置参数：run_daily_pipeline 的形参顺序一变，这里必须跟着改。
        # 参数错位不会报错，只会把 captcha_server 当成 recharge_cfg 之类静默跑歪。
        args=(platform, group_id, login_password, captcha_api_key, captcha_server,
              recharge_cfg),
        daemon=True,
    ).start()

    # 把实际生效的策略回给前端，让它能回显——用户配的和跑的是不是同一套，
    # 应当一眼可见，而不是只能去翻日志。
    return jsonify({"status": "started", "usable_cards": eligible,
                    "accounts": account_count,
                    "registerable_accounts": registerable_count,
                    "reusable_accounts": reusable_count,
                    "group_name": group['name'],
                    "amount_min": recharge_cfg.amount_min,
                    "amount_max": recharge_cfg.amount_max,
                    "balance_cap": recharge_cfg.balance_cap})


@api.route('/api/daily/subscribe/start', methods=['POST'])
def start_daily_subscribe_pipeline():
    """启动每日订阅任务：账号轮转——未注册先注册、已注册登录订阅，成功即换下一个账号，
    直到无可选卡 / 无待订阅账号 / 用户停止。与每日充值任务互斥（共用 is_running 闸门）。"""
    models = get_models()
    data = request.json or {}

    group_id = data.get('group_id')
    if not group_id:
        return jsonify({"error": "未指定卡池分组"}), 400
    group = models['card_group'].get_by_id(group_id)
    if not group:
        return jsonify({"error": "卡池分组不存在"}), 404

    platform = _req_platform(required=True)
    if not platform:
        return jsonify({"error": "未指定平台"}), 400

    # 闸门在拿到 platform 之后判，且只判这个平台（理由同每日充值端点）。
    state = get_ctx(platform)
    if state.is_running:
        return jsonify({"error": f"{platform} 已有任务在运行"}), 400

    captcha_api_key = data.get('captcha_api_key')
    # 订阅默认用 Multibot 解 hCaptcha；可传 captcha_server='2captcha.com' 切回
    captcha_server = data.get('captcha_server') or 'api.multibot.cloud'

    # 启动门：分组要有可选卡，且要有待订阅账号（判据同 _needing）
    # platform / exclude_used 必须显式传，理由同每日充值那个启动门。
    eligible = len(state._eligible_cards(group_id, exclude_used=False, platform=platform))
    if not eligible:
        return jsonify({"error": "该分组无可选卡（全部无效/过期或冷却中），无事可做"}), 400

    accts = models['account'].get_all(order_desc=False)
    pa_map = models['platform_account'].map_by_email(platform)
    account_count = sum(
        1 for a in accts
        if not is_identity_terminal(a.get('identity_status'))
        and not is_platform_terminal((pa_map.get(a['email']) or {}).get('status'))
    )
    if account_count == 0:
        return jsonify({"error": "无待订阅账号（身份或平台状态均已终态），无事可做"}), 400

    import threading
    threading.Thread(
        target=state.run_daily_subscribe_pipeline,
        args=(platform, group_id, captcha_api_key, captcha_server),
        daemon=True,
    ).start()

    return jsonify({"status": "started", "usable_cards": eligible,
                    "accounts": account_count, "group_name": group['name']})


# ==================== 代理 IP 池 ====================

def _mask_proxy(row):
    """列表展示用:凭据打码,只留首尾几位。"""
    def _m(s):
        s = s or ''
        return s if len(s) <= 4 else f"{s[:3]}***{s[-2:]}"
    return {
        'id': row['id'], 'host': row['host'], 'port': row['port'],
        'username': _m(row.get('username')), 'password': _m(row.get('password')),
        'status': row.get('status') or '', 'created_at': row.get('created_at'),
    }


@api.route('/api/proxies')
def get_proxies():
    models = get_models()
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('page_size', 50))
    rows, total = models['proxy'].get_all(page=page, page_size=page_size)
    return jsonify({"data": [_mask_proxy(r) for r in rows], "total": total,
                    "page": page, "page_size": page_size})


@api.route('/api/proxies/import', methods=['POST'])
def import_proxies():
    """粘贴导入:body {text} 每行一个 user:pass@host:port(兼容纯冒号)。"""
    models = get_models()
    text = (request.json or {}).get('text', '')
    if not text.strip():
        return jsonify({"error": "内容为空"}), 400
    added, skipped = models['proxy'].add_proxies(text)
    return jsonify({"added": added, "skipped": skipped, "total": models['proxy'].count()})


@api.route('/api/proxies/<int:proxy_id>', methods=['DELETE'])
def delete_proxy(proxy_id):
    models = get_models()
    models['proxy'].delete(proxy_id)
    return jsonify({"status": "ok"})


@api.route('/api/proxies/clear', methods=['POST'])
def clear_proxies():
    models = get_models()
    models['proxy'].clear()
    return jsonify({"status": "ok"})
