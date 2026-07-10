"""
信用卡 Excel 管理模块
负责解析 Excel 文件、生成模版、导出报告
"""

import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook

TEMPLATE_COLUMNS = [
    "card_number", "expiry_month", "expiry_year", "cvc",
    "first_name", "last_name", "country", "address",
    "address2", "city", "state", "zip", "company",
]

REQUIRED_FIELDS = [
    "card_number", "expiry_month", "expiry_year", "cvc",
    "first_name", "last_name", "country", "address", "city", "state", "zip",
]

STATUS_PENDING = "pending"
STATUS_SUCCESS = "success"
STATUS_FAILED = "failed"


def get_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def generate_template(save_path=None):
    if save_path is None:
        save_path = os.path.join(get_base_dir(), "credit_cards_template.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "CreditCards"

    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    example = [
        "4111111111111111", "12", "2026", "123",
        "John", "Doe", "United States",
        "123 Main St", "", "New York", "NY", "10001", ""
    ]
    for col_idx, val in enumerate(example, 1):
        ws.cell(row=2, column=col_idx, value=val)

    wb.save(save_path)
    return save_path


def parse_excel(file_path):
    errors = []

    if not os.path.exists(file_path):
        return [], [f"File not found: {file_path}"]

    try:
        wb = load_workbook(file_path, read_only=True)
    except Exception as e:
        return [], [f"Cannot open Excel file: {e}"]

    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    if len(rows) < 2:
        return [], ["Excel file is empty or has no data rows"]

    header = [str(cell).strip().lower() if cell else "" for cell in rows[0]]

    col_map = {}
    for idx, col_name in enumerate(header):
        normalized = col_name.replace(" ", "_")
        if normalized in TEMPLATE_COLUMNS:
            col_map[normalized] = idx
        elif normalized == "number":
            col_map["card_number"] = idx

    missing_cols = [f for f in REQUIRED_FIELDS if f not in col_map]
    if missing_cols:
        return [], [f"Missing required columns: {', '.join(missing_cols)}"]

    cards = []
    for row_idx, row in enumerate(rows[1:], start=2):
        card = {}
        row_errors = []

        for field in TEMPLATE_COLUMNS:
            if field in col_map:
                val = row[col_map[field]]
                card[field] = str(val).strip() if val is not None else ""
            else:
                card[field] = ""

        for field in REQUIRED_FIELDS:
            if not card.get(field):
                row_errors.append(f"Row {row_idx}: missing '{field}'")

        if row_errors:
            errors.extend(row_errors)
        else:
            cards.append({
                "number": card["card_number"],
                "expiry_month": card["expiry_month"],
                "expiry_year": card["expiry_year"],
                "cvc": card["cvc"],
                "first_name": card["first_name"],
                "last_name": card["last_name"],
                "country": card["country"],
                "address": card["address"],
                "address2": card.get("address2", ""),
                "city": card["city"],
                "state": card["state"],
                "zip": card["zip"],
                "company": card.get("company", ""),
            })

    wb.close()
    return cards, errors


def export_report(records, save_path=None):
    """导出绑卡报告 Excel"""
    if save_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = os.path.join(get_base_dir(), "data", f"bind_report_{timestamp}.xlsx")

    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "BindReport"

    headers = ["#", "Card (last 4)", "Status", "Bound To", "Error", "Time"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r.get("id", row_idx - 1))
        ws.cell(row=row_idx, column=2, value=f"****{r.get('card_display', '????')}")
        ws.cell(row=row_idx, column=3, value=r.get("status", ""))
        ws.cell(row=row_idx, column=4, value=r.get("bound_to_email", ""))
        ws.cell(row=row_idx, column=5, value=r.get("error", ""))
        ws.cell(row=row_idx, column=6, value=r.get("attempted_at", ""))

    wb.save(save_path)
    return save_path
