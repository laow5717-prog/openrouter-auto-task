"""账号导入：Excel ↔ accounts 表。

导入进来的账号一律落 identity_status='imported'，这是**待注册**的意思——每日任务的
补号环节（run_daily_pipeline._registerable_imported）只认这个状态，注册成功后由
_register_one_account 把它推进到 'registered'。导入时写别的状态会让账号要么永远
不被注册（'registered' 意味着「GitHub 已经开好了」），要么直接出局（终态）。

三个字段的用途各不相同，缺谁的后果也不同：

  email               主键。缺了整行跳过。
  email_password      邮箱密码。目前的收码链路（若安收信页）不需要它，留空也能注册；
                      存下来是为了将来换收码方式、或人工登录邮箱时不用另找。
  email_verify_link   若安收信链接。**这个才是注册能不能自动跑通的关键**：GitHub 注册
                      要邮箱验证码，靠它去收；指纹浏览器下每个账号都是全新环境，新设备
                      验证几乎必然触发。没有它的账号即便状态是 imported 也领不走——
                      _hotmail_for_account 取不到收码数据就把它过滤掉了，表现为「导入了
                      一堆账号但补号流程一个都不碰」。所以解析时缺它要单独报出来，
                      不能和「导入成功」混在一起。

列名匹配放宽到中英文各几种写法：这张表多半是从邮箱供应商那边拷过来的，列名五花八门，
为一个列名对不上就整批导不进去太蠢。
"""

from openpyxl import Workbook, load_workbook

from src.config import get_data_dir

# 规范字段 → 认得的列名（全部转小写去空格后比对）
_COLUMN_ALIASES = {
    'email': ('email', 'mail', '邮箱', '邮箱地址', '账号', '帐号', 'account'),
    'email_password': ('email_password', 'password', 'pwd', 'pass',
                       '邮箱密码', '密码'),
    'email_verify_link': ('email_verify_link', 'verify_link', 'link', 'url',
                          '邮箱认证链接', '认证链接', '收信链接', '验证链接', '链接'),
}

TEMPLATE_COLUMNS = ('邮箱', '邮箱密码', '邮箱认证链接')


def generate_template(save_path=None):
    """生成导入模版 xlsx，返回路径。"""
    if save_path is None:
        save_path = str(get_data_dir() / "accounts_template.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"

    for i, name in enumerate(TEMPLATE_COLUMNS, 1):
        ws.cell(row=1, column=i, value=name)

    # 示例行。写成一眼能看出「链接长什么样」的形状——第三列是三个字段里唯一
    # 无法凭直觉填对的，而它恰恰是自动注册能否跑通的关键。
    ws.cell(row=2, column=1, value="example001@hotmail.com")
    ws.cell(row=2, column=2, value="邮箱密码（可留空）")
    ws.cell(row=2, column=3, value="https://ruoanzhu.example/mail?e=example001@hotmail.com&p=xxx&h=xxx")

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 72

    wb.save(save_path)
    return save_path


def _header_map(header_row):
    """表头 → {规范字段: 列下标}。认不出的列忽略。"""
    got = {}
    for idx, raw in enumerate(header_row):
        name = str(raw or '').strip().lower().replace(' ', '').replace('_', '')
        for field, aliases in _COLUMN_ALIASES.items():
            if name in tuple(a.replace('_', '') for a in aliases):
                got.setdefault(field, idx)
                break
    return got


def parse_excel(path):
    """解析导入文件，返回 (rows, errors)。

    rows: [{email, email_password, email_verify_link}]，已按 email 去重（保留首次出现）。
    errors: 人类可读的问题列表，**不阻断**——能解析多少算多少，让用户一次看到全部问题，
            而不是修一个报一个。
    """
    errors = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return [], [f"无法打开文件: {e}"]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], ["文件为空或只有表头"]

    cols = _header_map(rows[0])
    if 'email' not in cols:
        return [], [f"未找到邮箱列（表头需含 {'/'.join(_COLUMN_ALIASES['email'][:4])} 之一）"]

    out = []
    seen = set()
    for n, row in enumerate(rows[1:], start=2):
        def _cell(field):
            i = cols.get(field)
            if i is None or i >= len(row):
                return ''
            return str(row[i]).strip() if row[i] is not None else ''

        email = _cell('email')
        if not email:
            continue                      # 空行，静默跳过——尾部空行太常见，报出来只是噪声
        if '@' not in email:
            errors.append(f"第 {n} 行：'{email}' 不像邮箱，已跳过")
            continue
        if email in seen:
            errors.append(f"第 {n} 行：{email} 在文件内重复，已跳过")
            continue
        seen.add(email)
        out.append({
            'email': email,
            'email_password': _cell('email_password'),
            'email_verify_link': _cell('email_verify_link'),
        })

    return out, errors


def import_rows(account_model, rows):
    """把解析结果写进 accounts，返回统计。

    一律 identity_status='imported'（待注册）。upsert 语义：已存在的账号只补空字段，
    不会把已有的密码/链接覆盖成空——重复导入同一份表是安全的。

    no_link 单独统计并回传：没有收码链接的账号虽然入了库，却**领不走**
    （_hotmail_for_account 取不到收码数据就过滤掉），不点出来的话用户会以为
    导入成功了、然后困惑于补号流程为什么不碰它们。
    """
    imported = 0
    no_link = []
    for r in rows:
        account_model.upsert(
            r['email'],
            email_password=r.get('email_password') or None,
            identity_status='imported',
            email_verify_link=r.get('email_verify_link') or None,
        )
        imported += 1
        if not (r.get('email_verify_link') or '').strip():
            no_link.append(r['email'])
    return {'imported': imported, 'no_link': no_link}
