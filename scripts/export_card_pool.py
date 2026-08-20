#!/usr/bin/env python3
"""把卡池导出成能直接被卡池导入功能吃进去的 xlsx。

列格式对齐 src/services/card.py 的 TEMPLATE_COLUMNS（13 列），目标机器上
「卡池 → 导入」选好卡组直接传即可。

产物写到 data/exports/card_pool_<时间戳>/：
  按卡组/<序号>_<组名>.xlsx   每个卡组一个文件，对应目标机器上建同名组
  可用卡_全部.xlsx            剔除过期/已废/已付后的干净卡，跨组合并
  全部卡.xlsx                 41k 行全量，含废卡，带卡组列
  有效卡.xlsx                 valid_cards 表
  README.md / manifest.json

标准 13 列之后还挂了几列 `_` 开头的状态（`_卡组` `_pool状态` `_opencode`
`_失败次数` `_3DS冷却至`）。导入端按列名匹配、不认识的列直接忽略，所以带着它们
不影响导入，但人能看出这张卡为什么被归到废卡里。

注意 xlsx **带不走状态**：导进目标机器的卡一律是全新的，之前攒的 invalid/paid
标记和 3DS 冷却窗口都会丢，那边会重新去试已经废掉的卡。要连状态一起搬就加
--with-sql，会额外产出一份 5 张表的 SQL 迁移包。

库正被服务写着的时候直接读会读到半截事务，所以先 VACUUM INTO 出一致性快照再导。
"""

import argparse
import json
import os
import re
import shutil
import sqlite3
import sys
from datetime import datetime

from openpyxl import Workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_DB = os.path.join(ROOT, "data", "openrouter_auto.db")

# 必须与 src/services/card.py 的 TEMPLATE_COLUMNS 一致，改那边记得改这里
CARD_COLUMNS = [
    "card_number", "expiry_month", "expiry_year", "cvc",
    "first_name", "last_name", "country", "address",
    "address2", "city", "state", "zip", "company",
]
# 导入端不认、纯给人看的诊断列
EXTRA_COLUMNS = ["_卡组", "_pool状态", "_opencode", "_失败次数", "_3DS冷却至"]

# SQL 迁移包用（--with-sql），顺序即导入顺序
SQL_TABLES = [
    "card_groups", "card_pool", "valid_cards",
    "card_payment_state", "card_platform_state",
]

PLATFORM = "opencode"


def snapshot(src_db, dst):
    """VACUUM INTO 出一份含 WAL 已提交内容的一致性副本。"""
    if os.path.exists(dst):
        os.remove(dst)
    conn = sqlite3.connect(f"file:{src_db}?mode=ro", uri=True)
    try:
        conn.execute("VACUUM INTO ?", (dst,))
    finally:
        conn.close()


def safe_name(s):
    """组名进文件名：干掉路径分隔符和 Windows 不收的字符。"""
    s = re.sub(r'[\\/:*?"<>|]', "_", str(s or "").strip())
    return (s or "未命名")[:40]


def write_sheet(rows, path, with_group_col):
    """写一个 xlsx。rows 是 sqlite3.Row 序列。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "CreditCards"

    extras = EXTRA_COLUMNS if with_group_col else EXTRA_COLUMNS[1:]
    header = CARD_COLUMNS + extras
    ws.append(header)

    for r in rows:
        line = []
        for c in CARD_COLUMNS:
            v = r[c]
            # 卡号/CVC/月份/年份一律按文本写。Excel 会把 16 位卡号转成科学计数法
            # 丢掉尾数，把 "06" 月吃成 6——这些字段回不来。
            line.append("" if v is None else str(v))
        if with_group_col:
            line.append(r["group_name"])
        line.extend([
            r["pool_status"] or "",
            r["plat_status"] or "",
            r["fail_streak"] if r["fail_streak"] else "",
            r["tds_until"] or "",
        ])
        ws.append(line)

    ws.freeze_panes = "A2"
    for col, w in zip("ABCDEFGHIJKLM", (20, 6, 6, 6, 14, 14, 16, 30, 12, 16, 14, 8, 24)):
        ws.column_dimensions[col].width = w
    wb.save(path)
    return ws.max_row - 1


BASE_SELECT = f"""
SELECT p.*, g.name AS group_name,
       IFNULL(p.status, '')       AS pool_status,
       s.status                   AS plat_status,
       ps.fail_streak             AS fail_streak,
       ps.tds_until               AS tds_until
FROM card_pool p
JOIN card_groups g ON g.id = p.group_id
LEFT JOIN card_platform_state s
       ON s.card_number = p.card_number AND s.platform = '{PLATFORM}'
LEFT JOIN card_payment_state ps
       ON ps.card_number = p.card_number AND ps.platform = '{PLATFORM}'
"""

# 干净卡：没过期、在目标平台上没留下 invalid/paid 记录
CLEAN_WHERE = "IFNULL(p.status,'') = '' AND s.status IS NULL"


def export_xlsx(conn, out_dir):
    made = []

    groups_dir = os.path.join(out_dir, "按卡组")
    os.makedirs(groups_dir, exist_ok=True)

    groups = list(conn.execute("SELECT id, name FROM card_groups ORDER BY id"))
    for gid, gname in groups:
        rows = list(conn.execute(BASE_SELECT + " WHERE p.group_id = ? ORDER BY p.id", (gid,)))
        if not rows:
            continue
        clean = sum(1 for r in rows if not r["pool_status"] and not r["plat_status"])
        fn = f"{gid:02d}_{safe_name(gname)}_{len(rows)}张_可用{clean}.xlsx"
        path = os.path.join(groups_dir, fn)
        n = write_sheet(rows, path, with_group_col=False)
        made.append((os.path.join("按卡组", fn), n))

    rows = list(conn.execute(BASE_SELECT + f" WHERE {CLEAN_WHERE} ORDER BY p.group_id, p.id"))
    fn = f"可用卡_全部_{len(rows)}张.xlsx"
    made.append((fn, write_sheet(rows, os.path.join(out_dir, fn), True)))

    rows = list(conn.execute(BASE_SELECT + " ORDER BY p.group_id, p.id"))
    fn = f"全部卡_{len(rows)}张.xlsx"
    made.append((fn, write_sheet(rows, os.path.join(out_dir, fn), True)))

    made.append(export_valid(conn, out_dir))
    return made


def export_valid(conn, out_dir):
    """valid_cards 单独一张表：列名不同，且多了 platform / validated_at。"""
    rows = list(conn.execute(
        "SELECT * FROM valid_cards ORDER BY platform, validated_at"))
    fn = f"有效卡_{len(rows)}张.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ValidCards"
    ws.append(CARD_COLUMNS + ["_platform", "_source_type", "_source_email", "_validated_at"])
    for r in rows:
        line = ["" if r[c] is None else str(r[c]) for c in CARD_COLUMNS]
        line += [r["platform"], r["source_type"], r["source_email"], r["validated_at"]]
        ws.append(line)
    ws.freeze_panes = "A2"
    for col, w in zip("ABCDEFGHIJKLM", (20, 6, 6, 6, 14, 14, 16, 30, 12, 16, 14, 8, 24)):
        ws.column_dimensions[col].width = w
    wb.save(os.path.join(out_dir, fn))
    return (fn, len(rows))


# ---------------- SQL 迁移包（--with-sql，连状态一起搬） ----------------

def sql_literal(v):
    if v is None:
        return "NULL"
    if isinstance(v, (int, float)):
        return repr(v)
    if isinstance(v, bytes):
        return "X'" + v.hex() + "'"
    return "'" + str(v).replace("'", "''") + "'"


def dump_sql(conn, out_path):
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("-- 卡池迁移包 / generated by scripts/export_card_pool.py\n")
        f.write("PRAGMA foreign_keys=OFF;\nBEGIN TRANSACTION;\n\n")
        for t in SQL_TABLES:
            row = conn.execute(
                "SELECT sql FROM sqlite_master WHERE type='table' AND name=?", (t,)
            ).fetchone()
            if not row:
                print(f"  [skip] 源库没有表 {t}", file=sys.stderr)
                continue
            f.write(f"-- ===== {t} =====\nDROP TABLE IF EXISTS {t};\n")
            f.write(row[0].rstrip().rstrip(";") + ";\n")
            cols = [c[1] for c in conn.execute(f"PRAGMA table_info({t})")]
            collist = ", ".join(f'"{c}"' for c in cols)
            n, batch = 0, []
            for r in conn.execute(f"SELECT * FROM {t}"):
                batch.append("(" + ", ".join(sql_literal(v) for v in r) + ")")
                n += 1
                if len(batch) >= 500:
                    f.write(f"INSERT INTO {t} ({collist}) VALUES\n" + ",\n".join(batch) + ";\n")
                    batch = []
            if batch:
                f.write(f"INSERT INTO {t} ({collist}) VALUES\n" + ",\n".join(batch) + ";\n")
            for (isql,) in conn.execute(
                "SELECT sql FROM sqlite_master WHERE type='index' AND tbl_name=? "
                "AND sql IS NOT NULL", (t,)
            ):
                f.write(isql.rstrip().rstrip(";") + ";\n")
            f.write(f"-- {t}: {n} 行\n\n")
        f.write("COMMIT;\nPRAGMA foreign_keys=ON;\n")


def main():
    ap = argparse.ArgumentParser(description="导出卡池 xlsx（卡池导入功能可直接使用）")
    ap.add_argument("--db", default=DEFAULT_DB, help=f"源库，默认 {DEFAULT_DB}")
    ap.add_argument("--out", default=None, help="输出目录，默认 data/exports/card_pool_<时间戳>")
    ap.add_argument("--with-sql", action="store_true",
                    help="额外产出 SQL 迁移包，把 invalid/paid/3DS 冷却等状态一起搬过去")
    args = ap.parse_args()

    if not os.path.exists(args.db):
        sys.exit(f"源库不存在：{args.db}")

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_dir = args.out or os.path.join(ROOT, "data", "exports", f"card_pool_{stamp}")
    os.makedirs(out_dir, exist_ok=True)

    snap = os.path.join(out_dir, ".snapshot.db")
    print(f"[1/3] 快照 {args.db}")
    snapshot(args.db, snap)

    conn = sqlite3.connect(f"file:{snap}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    conn.text_factory = str
    try:
        print("[2/3] 导出 xlsx")
        made = export_xlsx(conn, out_dir)

        if args.with_sql:
            print("      额外导出 cards.sql（含状态）")
            dump_sql(conn, os.path.join(out_dir, "cards.sql"))
            shutil.copy2(
                os.path.join(ROOT, "scripts", "import_card_pool.py"),
                os.path.join(out_dir, "import_card_pool.py"),
            )

        print("[3/3] 生成 README / manifest")
        stats = build_stats(conn)
    finally:
        conn.close()
    os.remove(snap)

    manifest = {
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "source_db": os.path.abspath(args.db),
        "columns": CARD_COLUMNS,
        "files": [{"name": n, "rows": c} for n, c in made],
        "stats": stats,
        "with_sql": args.with_sql,
    }
    with open(os.path.join(out_dir, "manifest.json"), "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    write_readme(out_dir, manifest)

    print(f"\n完成：{out_dir}")
    for n, c in made:
        print(f"  {n:<44s} {c:>7,} 行")


def build_stats(conn):
    per_group = [
        dict(zip(("id", "name", "total", "expired", "invalid", "paid", "clean"), r))
        for r in conn.execute(f"""
            SELECT g.id, g.name, COUNT(*),
              SUM(CASE WHEN IFNULL(p.status,'')='expired' THEN 1 ELSE 0 END),
              SUM(CASE WHEN s.status='invalid' THEN 1 ELSE 0 END),
              SUM(CASE WHEN s.status='paid' THEN 1 ELSE 0 END),
              SUM(CASE WHEN {CLEAN_WHERE} THEN 1 ELSE 0 END)
            FROM card_pool p JOIN card_groups g ON g.id=p.group_id
            LEFT JOIN card_platform_state s
                   ON s.card_number=p.card_number AND s.platform='{PLATFORM}'
            GROUP BY g.id ORDER BY g.id""")
    ]
    return {
        "platform": PLATFORM,
        "per_group": per_group,
        "total": sum(g["total"] for g in per_group),
        "clean": sum(g["clean"] for g in per_group),
    }


def write_readme(out_dir, m):
    s = m["stats"]
    g_lines = "\n".join(
        f"| {g['id']} | {g['name']} | {g['total']:,} | {g['clean']:,} | "
        f"{g['invalid']:,} | {g['paid']} | {g['expired']} |"
        for g in s["per_group"]
    )
    f_lines = "\n".join(f"| `{x['name']}` | {x['rows']:,} |" for x in m["files"])
    sql_note = (
        "\n## 连状态一起搬（cards.sql）\n\n"
        "本次带了 `cards.sql` + `import_card_pool.py`。xlsx 带不走状态，这份能：\n\n"
        "```bash\npython3 import_card_pool.py --db /path/to/data/openrouter_auto.db --dry-run\n"
        "python3 import_card_pool.py --db /path/to/data/openrouter_auto.db\n```\n\n"
        "卡组按名称匹配、`group_id` 自动重映射，卡片按 `(card_number, group_id)` 去重，"
        "重复执行安全。\n"
        if m["with_sql"] else
        "\n## 想连状态一起搬\n\n"
        "xlsx 带不走 invalid / paid / 3DS 冷却。要保留就重跑一次导出加 `--with-sql`，"
        "会多出一份 `cards.sql` 和 `import_card_pool.py`。\n"
    )
    body = f"""# 卡池导出（xlsx）

导出时间：{m['exported_at']}
源库：`{m['source_db']}`

> 内含真实卡号 / CVC / 持卡人信息。**不要提交到 git，不要走 IM 或网盘明文传输。**

## 文件

| 文件 | 行数 |
| --- | ---: |
{f_lines}

- **`按卡组/`** —— 一组一个文件。目标机器上按同名建卡组，再逐个导入，卡组结构原样保留。
- **`可用卡_全部.xlsx`** —— 剔除过期 / 已废 / 已付后的干净卡，跨组合并成一个文件。
  只想快速开跑、不在乎分组的话导这个。
- **`全部卡.xlsx`** —— 含废卡的 41k 全量，留档用。
- **`有效卡.xlsx`** —— `valid_cards` 表，验证通过的卡。

## 列格式

前 13 列就是卡池导入模版的列，和 `src/services/card.py` 的 `TEMPLATE_COLUMNS` 一致：

```
card_number, expiry_month, expiry_year, cvc, first_name, last_name,
country, address, address2, city, state, zip, company
```

其中 `address2` 和 `company` 可空，其余 11 列必填。

后面 `_` 开头的几列（`_卡组` `_pool状态` `_opencode` `_失败次数` `_3DS冷却至`）
是给人看的诊断信息。导入端按列名匹配、不认识的列直接忽略，所以留着不影响导入。

所有单元格按文本写入，避免 Excel 把 16 位卡号转成科学计数法丢尾数、把月份 `06` 吃成 `6`。

## 各卡组（平台口径：{s['platform']}）

| id | 卡组 | 总数 | 可用 | 已废 | 已付 | 过期 |
| ---: | --- | ---: | ---: | ---: | ---: | ---: |
{g_lines}

合计 {s['total']:,} 张，其中干净可用 {s['clean']:,} 张。

## 导入到目标机器

卡池页面 → 新建卡组 → 导入，选对应 xlsx 即可。

**xlsx 只带卡，不带状态。** 导进去的卡一律是全新的：这边攒下的
{s['total'] - s['clean']:,} 张废卡记录（invalid / paid / 过期）和 3DS 冷却窗口都不会跟过去。
所以直接导 `全部卡.xlsx` 会让目标机器把已经废掉的卡重新试一遍。
一般导 `可用卡_全部.xlsx` 或按卡组导，把废卡挡在外面。
{sql_note}"""
    with open(os.path.join(out_dir, "README.md"), "w", encoding="utf-8") as f:
        f.write(body)


if __name__ == "__main__":
    main()
