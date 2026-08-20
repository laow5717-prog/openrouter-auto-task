#!/usr/bin/env python3
"""把邮箱供应商给的 `邮箱----密码----收信链接` 单列文件转成账号导入模版格式。

供应商发过来的表基本都是一列塞满、`----` 分隔、没有表头（`底料/2D.txt` 也是这个
形状），而 src/services/account_import.py 认的是「邮箱 / 邮箱密码 / 邮箱认证链接」
三列带表头。这个脚本做的就是这一层翻译。

    python3 scripts/convert_email_import.py "~/Downloads/20 (2).xlsx"
    python3 scripts/convert_email_import.py 源.txt -o 输出.xlsx

输入支持 .xlsx / .txt / .csv。分隔符默认 `----`，用 --sep 换。
"""

import argparse
import os
import sys

from openpyxl import Workbook, load_workbook

# 与 account_import.TEMPLATE_COLUMNS 保持一致，改那边记得改这里
HEADER = ("邮箱", "邮箱密码", "邮箱认证链接")


def read_lines(path):
    """把输入文件读成一行行字符串。xlsx 走 openpyxl，其余按文本读。"""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        out = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                # 供应商的表偶尔会把整条塞在第一列，偶尔又拆成几列；
                # 统一先按非空单元格拼回一行，后面再按分隔符切。
                cells = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if cells:
                    out.append(cells[0] if len(cells) == 1 else "----".join(cells))
        wb.close()
        return out
    with open(path, encoding="utf-8-sig", errors="replace") as f:
        return [ln.strip() for ln in f if ln.strip()]


def looks_like_header(parts):
    """首行是表头就别当数据用掉。"""
    joined = "".join(parts).lower()
    return "@" not in joined and any(
        k in joined for k in ("邮箱", "mail", "账号", "帐号", "密码", "password")
    )


def parse(lines, sep):
    rows, errors = [], []
    seen = set()
    for n, line in enumerate(lines, start=1):
        parts = [p.strip() for p in line.split(sep)]
        if n == 1 and looks_like_header(parts):
            continue
        email = parts[0] if parts else ""
        if not email:
            continue
        if "@" not in email:
            errors.append(f"第 {n} 行：'{email[:40]}' 不像邮箱，已跳过")
            continue
        pwd = parts[1] if len(parts) > 1 else ""
        link = parts[2] if len(parts) > 2 else ""

        # 链接里的 e= 就是这条记录的邮箱。两者对不上说明供应商的表串行了，
        # 导进去的后果是拿 A 的链接去收 B 的验证码，注册永远卡在验证这步。
        if link and "e=" in link:
            in_link = link.split("e=", 1)[1].split("&", 1)[0]
            if in_link.lower() != email.lower():
                errors.append(f"第 {n} 行：{email} 与链接内的 {in_link} 不一致，已跳过")
                continue
        if email.lower() in seen:
            errors.append(f"第 {n} 行：{email} 重复，已跳过")
            continue
        seen.add(email.lower())

        if not link:
            errors.append(f"第 {n} 行：{email} 缺收信链接，导入后无法自动注册")
        rows.append((email, pwd, link))
    return rows, errors


def write_xlsx(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounts"
    for i, name in enumerate(HEADER, 1):
        ws.cell(row=1, column=i, value=name)
    for r, (email, pwd, link) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=email)
        ws.cell(row=r, column=2, value=pwd)
        ws.cell(row=r, column=3, value=link)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 78
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="邮箱供应商文件 → 账号导入模版格式")
    ap.add_argument("src", help="源文件 .xlsx / .txt / .csv")
    ap.add_argument("-o", "--out", help="输出路径，默认源文件同目录加 _导入格式.xlsx")
    ap.add_argument("--sep", default="----", help="字段分隔符，默认 ----")
    args = ap.parse_args()

    src = os.path.expanduser(args.src)
    if not os.path.exists(src):
        sys.exit(f"源文件不存在：{src}")

    out = args.out or os.path.join(
        os.path.dirname(os.path.abspath(src)),
        os.path.splitext(os.path.basename(src))[0] + "_导入格式.xlsx",
    )
    out = os.path.expanduser(out)

    lines = read_lines(src)
    rows, errors = parse(lines, args.sep)
    if not rows:
        sys.exit("没解析出任何账号" + (f"：{errors[0]}" if errors else ""))

    write_xlsx(rows, out)

    print(f"源文件   {src}")
    print(f"读入     {len(lines)} 行")
    print(f"输出     {out}")
    print(f"账号     {len(rows)} 个，带收信链接 {sum(1 for r in rows if r[2])} 个")
    if errors:
        print(f"\n{len(errors)} 条问题：")
        for e in errors[:20]:
            print(f"  {e}")
        if len(errors) > 20:
            print(f"  ... 另有 {len(errors) - 20} 条")


if __name__ == "__main__":
    main()
