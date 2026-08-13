"""把「底料」目录里各种格式的卡表合并成 credit_cards_template.xlsx 的格式。

底料有两种列头方言（英文 CardNumber/... 与中文 卡号/...），每个文件还有多个
sheet（Orders/有效/2D/绑/重复/...），部分文件带密码。这里统一：识别方言 → 映射到
模版 13 列 → 清洗校验 → 跨文件跨 sheet 按卡号去重 → 写单 sheet 输出。

用法:
    python scripts/merge_card_sources.py [--src 底料] [--out 底料/合并结果.xlsx]
    python scripts/merge_card_sources.py --exclude-countries MX,CN,JP,BR
    python scripts/merge_card_sources.py --no-exclude-countries
"""

import argparse
import io
import json
import os
import re
import sys
from collections import Counter

import openpyxl

# 加密底料的密码（file 报 CDFV2 Encrypted 的那几个）
PASSWORDS = ["5.0", "303", "VelvetSweatshop", ""]

TEMPLATE_COLUMNS = [
    "card_number", "expiry_month", "expiry_year", "cvc",
    "first_name", "last_name", "country", "address",
    "address2", "city", "state", "zip", "company",
]

# 卡本身的核心字段，缺了这些这条记录没有意义
REQUIRED_FIELDS = ["card_number", "expiry_month", "expiry_year", "cvc"]

# 账单信息，缺失不判废，但用来衡量记录完整度（同卡号去重时保留更全的那条）
BILLING_FIELDS = [
    "first_name", "last_name", "country", "address", "address2",
    "city", "state", "zip",
]

# --require-address 时才要求补齐的字段
ADDRESS_REQUIRED = ["first_name", "last_name", "country", "address", "city", "state", "zip"]

# 这几个国家的卡在下游支付里不可用，之前已从 card_pool 里清过一轮，
# 默认在合并阶段就剔掉，免得下次导入又混进来。
# 用 --exclude-countries 换一套，用 --no-exclude-countries 关掉。
# 注意只对能规范成 ISO2 的卡生效：国家列认不出（留空）的卡照收不误。
DEFAULT_EXCLUDED_COUNTRIES = ("MX", "CN", "JP")

# 底料有好几套列头方言（英文 CardNumber/…、账单前缀的 账单国家/账单地址1/…、
# 以及裸中文 国家/地址1/城市/州/邮编/…），全部映射到模版字段。
# 注意 姓/名 两列：底料里实际是 姓=名字、名=姓氏（对照同表「持卡人」可验证），
# 所以按真实取值映射，不按字面意思。
ALIASES = {
    "card_number": ("cardnumber", "卡号", "number", "card_number"),
    "cvc": ("cvv", "cvc", "cvv2"),
    "first_name": ("firstname", "名字", "first_name", "姓"),
    "last_name": ("lastname", "姓氏", "last_name", "名"),
    "country": ("country", "账单国家", "国家", "国家名称", "国家代码"),
    "address": ("street1", "账单地址1", "address", "地址1", "地址", "账单地址"),
    "address2": ("street2", "账单地址2", "address2", "地址2"),
    "city": ("city", "账单城市", "城市"),
    "state": ("state", "账单省", "州", "省", "省份"),
    "zip": ("zipcode", "账单邮编", "zip", "postalcode", "邮编", "邮政编码"),
}
EXPIRY_ALIASES = ("expirydate", "有效期", "expiry", "exp", "过期时间", "日期", "到期日")

# 单列姓名，按优先级取第一个命中的列，拆成 first/last
FULLNAME_ALIASES = ("持卡人", "持卡人姓名", "卡名", "cardholder", "cardname", "姓名", "name")

# 发卡行国家，只在没有账单国家列时兜底
COUNTRY_FALLBACK_ALIASES = ("卡国家", "卡片国家", "发行国家", "发卡国家", "卡片发行国家")

STATE_ABBR = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "district of columbia": "DC", "florida": "FL", "georgia": "GA", "hawaii": "HI",
    "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM", "new york": "NY",
    "north carolina": "NC", "north dakota": "ND", "ohio": "OH", "oklahoma": "OK",
    "oregon": "OR", "pennsylvania": "PA", "rhode island": "RI",
    "south carolina": "SC", "south dakota": "SD", "tennessee": "TN", "texas": "TX",
    "utah": "UT", "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
    "puerto rico": "PR", "guam": "GU", "virgin islands": "VI",
    # 旧式/口语缩写，底料里成片出现（Fla. / OKLA / Calif.）
    "fla": "FL", "calif": "CA", "okla": "OK", "penna": "PA", "penn": "PA",
    "mass": "MA", "conn": "CT", "tenn": "TN", "wash": "WA", "mich": "MI",
    "minn": "MN", "miss": "MS", "wis": "WI", "wisc": "WI", "ariz": "AZ",
    "ark": "AR", "colo": "CO", "ill": "IL", "kans": "KS", "mont": "MT",
    "nebr": "NE", "nev": "NV", "ore": "OR", "oreg": "OR", "tex": "TX",
    "wyo": "WY", "n mex": "NM", "n car": "NC", "s car": "SC",
    # 底料里混进来的外语/变体写法
    "washington dc": "DC", "washington d.c.": "DC", "pennsylvanie": "PA",
    "californie": "CA", "floride": "FL", "géorgie": "GA", "조지아 주": "GA",
}

US_STATES = set(STATE_ABBR.values())

COUNTRY_ABBR = {
    "united states": "US", "usa": "US", "us": "US",
    "united states of america": "US",
    "canada": "CA", "united kingdom": "GB", "uk": "GB", "great britain": "GB",
    "mexico": "MX", "méxico": "MX", "australia": "AU", "brazil": "BR",
    "brasil": "BR", "india": "IN", "germany": "DE", "france": "FR",
    "spain": "ES", "italy": "IT", "japan": "JP", "netherlands": "NL",
    "singapore": "SG", "south korea": "KR", "korea": "KR", "new zealand": "NZ",
    "ireland": "IE", "switzerland": "CH", "sweden": "SE", "poland": "PL",
    # 底料里「卡国家」列常见的中文写法
    "美国": "US", "加拿大": "CA", "英国": "GB", "澳大利亚": "AU", "日本": "JP",
    "德国": "DE", "法国": "FR", "墨西哥": "MX", "巴西": "BR", "新加坡": "SG",
    "韩国": "KR", "意大利": "IT", "西班牙": "ES", "荷兰": "NL", "印度": "IN",
    "中国": "CN", "中国大陆": "CN", "香港": "HK", "台湾": "TW",
}


def compact(text):
    return "".join(ch for ch in str(text or "").lower() if ch.isalnum())


def build_country_index():
    """国家名 → ISO2。有 pycountry 就用全量表，没有就退回上面的手写表。"""
    index = {}
    try:
        import pycountry
    except ImportError:
        pass
    else:
        for c in pycountry.countries:
            names = (c.name, getattr(c, "official_name", None), getattr(c, "common_name", None))
            for name in filter(None, names):
                index[compact(name)] = c.alpha_2
            index[c.alpha_2.lower()] = c.alpha_2
            index[c.alpha_3.lower()] = c.alpha_2
    for name, code in COUNTRY_ABBR.items():
        index[compact(name)] = code
    index["kosovo"] = "XK"          # 不在 ISO 3166 里，底料出现过
    return index


COUNTRY_INDEX = build_country_index()


def normalize_country(raw):
    """底料的国家列不可信（混进过邮编、User-Agent、卡组织），认不出就返回空。"""
    text = str(raw or "").strip()
    if not text:
        return ""
    for candidate in (text, text.split(",")[0], text.split("(")[0]):
        code = COUNTRY_INDEX.get(compact(candidate))
        if code:
            return code
    return ""


def clean_text(value):
    """列错位会把 User-Agent 这类长串灌进地址/姓名列，宁可留空也别写进结果。"""
    text = str(value or "").strip()
    if "Mozilla/" in text or "AppleWebKit" in text or len(text) > 120:
        return ""
    return text


def split_name(full):
    """'Brynn Munoz' → ('Brynn', 'Munoz')；只有一个词时姓名同值，总比丢掉整行强。"""
    parts = str(full or "").split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], parts[0]
    return parts[0], " ".join(parts[1:])


def make_card(number, month, year, cvc, first, last,
              country_candidates, address, address2, city, state, zip_raw):
    """把各路来源的原始字段统一成模版格式，顺带清洗州/邮编/国家。"""
    state = str(state or "").strip()
    state = STATE_ABBR.get(state.lower().replace(".", "").strip(), state.rstrip("."))
    if len(state) == 2 and state.isalpha():
        state = state.upper()          # 底料里大量写成 Fl / Tx / fl

    country = ""
    for candidate in country_candidates:
        country = normalize_country(candidate)
        if country:
            break
    # 国家列认不出（或压根没有）时，州是美国州缩写就足以判定 US
    if not country and state.upper() in US_STATES:
        country = "US"

    # 邮编有 ZIP+4（21228-5324）写法，直接抽数字会拼成 9 位；
    # 另有一批被 Excel 当数字存过，前导零被吃掉（09256 → 9256）
    zip_raw = str(zip_raw or "").strip()
    zip_value = digits(zip_raw)
    if len(zip_value) == 9:
        zip_value = zip_value[:5]
    elif country == "US" and 3 <= len(zip_value) <= 4 and zip_value == zip_raw:
        zip_value = zip_value.zfill(5)
    zip_value = zip_value or zip_raw

    # 州名不会带数字，带了就是列错位串进来的电话号 / UA；邮编也有个大致形态。
    # 这两格宁可留空，但卡本身还是留着。
    state = clean_text(state)
    if len(state) > 40 or any(ch.isdigit() for ch in state):
        state = ""
    zip_value = clean_text(zip_value)
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9 -]{2,9}", zip_value):
        zip_value = ""

    # 美国的州/邮编形态固定，可以卡得更死
    if country == "US":
        if state not in US_STATES:
            state = ""
        if not (len(zip_value) == 5 and zip_value.isdigit()):
            zip_value = ""

    return {
        "card_number": number,
        "expiry_month": month or "",
        "expiry_year": year or "",
        "cvc": digits(cvc),
        "first_name": clean_text(first),
        "last_name": clean_text(last),
        "country": country,
        "address": clean_text(address),
        "address2": clean_text(address2),
        "city": clean_text(city),
        "state": state,
        "zip": zip_value,
        "company": "",
    }


def completeness(card):
    """先比必需的账单字段齐了几个，再比总字段数，免得可选的 address2 顶掉 state。"""
    return (
        sum(1 for f in ADDRESS_REQUIRED if card.get(f)),
        sum(1 for f in BILLING_FIELDS if card.get(f)),
    )


def norm_header(cell):
    return str(cell or "").strip().lower().replace(" ", "").replace("_", "")


def digits(text):
    return "".join(ch for ch in str(text or "") if ch.isdigit())


def luhn_ok(number):
    total, alt = 0, False
    for ch in reversed(number):
        d = ord(ch) - 48
        if alt:
            d *= 2
            if d > 9:
                d -= 9
        total += d
        alt = not alt
    return total % 10 == 0


def parse_expiry(raw):
    """'11/26' / '05/2030' / '1126' / datetime → ('11', '2026')，失败返回 (None, None)。"""
    if raw is None:
        return None, None
    if hasattr(raw, "month") and hasattr(raw, "year"):
        return f"{raw.month:02d}", str(raw.year)

    # 有的表把有效期存成了 Excel 日期序列号（46112 → 2026-03-31）
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        if 40000 <= raw <= 60000:
            import datetime
            d = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(raw))
            return f"{d.month:02d}", str(d.year)

    text = str(raw).strip()
    parts = [p for p in text.replace("-", "/").replace(".", "/").split("/") if p.strip()]
    if len(parts) >= 2:
        a, b = digits(parts[0]), digits(parts[1])
        # 有的表写成 2026/11
        if len(a) == 4:
            a, b = b, a
        month, year = a, b
    else:
        d = digits(text)
        if len(d) == 4:      # MMYY
            month, year = d[:2], d[2:]
        elif len(d) == 6:    # MMYYYY
            month, year = d[:2], d[2:]
        else:
            return None, None

    if not month or not year:
        return None, None
    try:
        m = int(month)
    except ValueError:
        return None, None
    if not 1 <= m <= 12:
        return None, None

    if len(year) == 2:
        year = "20" + year
    if len(year) != 4:
        return None, None
    return f"{m:02d}", year


def load_legacy_xls(path):
    """WPS 存的旧版 BIFF .xls，openpyxl 读不了，用 xlrd 转成内存 workbook。"""
    import xlrd

    book = xlrd.open_workbook(path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sh in book.sheets():
        ws = wb.create_sheet((sh.name or "Sheet")[:31])
        for r in range(sh.nrows):
            row = []
            for value, ctype in zip(sh.row_values(r), sh.row_types(r)):
                # 数值型卡号别留成浮点，否则会被写成科学计数
                if ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
                    value = int(value)
                row.append(None if value == "" else value)
            ws.append(row)
    return wb


def load_workbook_any(path):
    """普通 xlsx 直接读；旧版 .xls 走 xlrd；加密的挨个试 PASSWORDS。"""
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        pass

    if path.lower().endswith(".xls"):
        try:
            return load_legacy_xls(path)
        except Exception:
            pass

    import msoffcrypto

    last = None
    for pw in PASSWORDS:
        try:
            with open(path, "rb") as fh:
                office = msoffcrypto.OfficeFile(fh)
                office.load_key(password=pw)
                buf = io.BytesIO()
                office.decrypt(buf)
            return openpyxl.load_workbook(buf, read_only=True, data_only=True)
        except Exception as e:
            last = e
    raise RuntimeError(f"无法打开（密码都不对）: {last}")


def build_col_map(header, sample_rows):
    """表头 → {模版字段: 列下标}，另含 __expiry__。认不出的 sheet 返回 None。

    底料的表头并不可靠（见 7.27猫头鹰/Orders，卡号列的表头被覆盖成 '7714'），
    所以表头认不出卡号/有效期时，退回按列内值的形态推断。
    """
    normalized = [norm_header(c) for c in header]
    col_map = {}
    for field, keys in ALIASES.items():
        for idx, name in enumerate(normalized):
            if name in keys and field not in col_map:
                col_map[field] = idx
    for idx, name in enumerate(normalized):
        if name in EXPIRY_ALIASES:
            col_map["__expiry__"] = idx
            break

    # 单列姓名和发卡行国家按别名优先级取，不按列序
    for alias in FULLNAME_ALIASES:
        if alias in normalized:
            col_map["__fullname__"] = normalized.index(alias)
            break
    for alias in COUNTRY_FALLBACK_ALIASES:
        if alias in normalized:
            col_map["__country2__"] = normalized.index(alias)
            break

    # 只有确定这是一张卡表（认出足够多字段）才允许按形态补列，避免误判杂表
    if len(col_map) < 5 or not sample_rows:
        return col_map if {"card_number", "__expiry__"} <= col_map.keys() else None

    taken = set(col_map.values())
    width = max(len(r) for r in sample_rows)

    if "card_number" not in col_map:
        for idx in range(width):
            if idx in taken:
                continue
            vals = [digits(r[idx]) for r in sample_rows if idx < len(r) and r[idx] is not None]
            hits = [v for v in vals if 13 <= len(v) <= 19 and luhn_ok(v)]
            if vals and len(hits) >= max(2, len(vals) * 0.6):
                col_map["card_number"] = idx
                taken.add(idx)
                break

    if "__expiry__" not in col_map:
        for idx in range(width):
            if idx in taken:
                continue
            vals = [r[idx] for r in sample_rows if idx < len(r) and r[idx] is not None]
            hits = [v for v in vals if parse_expiry(v) != (None, None) and "/" in str(v)]
            if vals and len(hits) >= max(2, len(vals) * 0.6):
                col_map["__expiry__"] = idx
                break

    if "card_number" not in col_map or "__expiry__" not in col_map:
        return None
    return col_map


# 有的表把整条卡记录以 JSON 串塞进单个单元格（数据归总/局外人/外卖那批）
JSON_ALIASES = {
    "card_number": ("cardno", "cardnumber", "number"),
    "cvc": ("cardcvv", "cvv", "cvc", "cardcvc"),
    "address": ("street1", "address", "addressline1", "street", "address1"),
    "address2": ("street2", "address2", "addressline2"),
    "city": ("city",),
    "state": ("state", "province", "region"),
    "zip": ("zipcode", "zip", "postalcode", "postcode"),
}
JSON_COUNTRY_KEYS = ("countryname", "country", "countrycode", "issuercountry")
JSON_NAME_KEYS = ("cardname", "cardholder", "name", "holdername")
JSON_MONTH_KEYS = ("cardmonth", "expmonth", "expirymonth", "month")
JSON_YEAR_KEYS = ("cardyear", "expyear", "expiryyear", "year")


def card_from_json(item):
    """一个 JSON 卡对象 → 模版格式；缺卡号就返回 None。"""
    fields = {str(k).lower(): v for k, v in item.items()}

    def pick(keys):
        for k in keys:
            if fields.get(k) not in (None, ""):
                return fields[k]
        return ""

    number = digits(pick(JSON_ALIASES["card_number"]))
    if not number:
        return None

    month, year = parse_expiry(f"{pick(JSON_MONTH_KEYS)}/{pick(JSON_YEAR_KEYS)}")
    first, last = split_name(pick(JSON_NAME_KEYS))
    return make_card(
        number, month, year, pick(JSON_ALIASES["cvc"]), first, last,
        [pick(JSON_COUNTRY_KEYS)],
        pick(JSON_ALIASES["address"]), pick(JSON_ALIASES["address2"]),
        pick(JSON_ALIASES["city"]), pick(JSON_ALIASES["state"]),
        pick(JSON_ALIASES["zip"]),
    )


def extract_json_cards(rows, stats):
    """扫出单元格里的 JSON 卡记录。这类 sheet 的表头往往认不出，得单独捞。"""
    out = []
    for row in rows:
        for value in row:
            if not isinstance(value, str) or '"cardno"' not in value.lower():
                continue
            try:
                obj = json.loads(value.strip())
            except ValueError:
                stats["json_unparsable"] += 1
                continue
            for item in (obj if isinstance(obj, list) else [obj]):
                if not isinstance(item, dict):
                    continue
                card = card_from_json(item)
                if card:
                    out.append(card)
                    stats["json_card"] += 1
    return out


def extract_rows(ws, stats):
    """从一个 sheet 里抽出模版格式的行。表头不认识就只捞 JSON 卡。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    out = extract_json_cards(rows, stats)
    header, body = rows[0], rows[1:]

    col_map = build_col_map(header, body[:30])
    if col_map is None:
        stats["sheet_skipped"] += 1
        return out

    for row in body:
        def cell(field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return ""
            val = row[idx]
            if val is None:
                return ""
            # 整数存成了浮点时 str() 会带上 .0，抽数字会把邮编 7308.0 变成 73080
            if isinstance(val, float) and val.is_integer():
                val = int(val)
            return str(val).strip()

        number = digits(cell("card_number"))
        if not number:
            continue

        raw_expiry = row[col_map["__expiry__"]] if col_map["__expiry__"] < len(row) else None
        month, year = parse_expiry(raw_expiry)

        # 同一个 sheet 里可能拼接了两段列序不同的数据（7.25猫头鹰(2)/Orders 从第
        # 333 行起 有效期/CVV 互换）。表头说了不算，按值形态逐行纠正。
        if month is None:
            raw_cvc = row[col_map["cvc"]] if "cvc" in col_map and col_map["cvc"] < len(row) else None
            swapped_m, swapped_y = parse_expiry(raw_cvc)
            if swapped_m is not None:
                month, year = swapped_m, swapped_y
                col_map = dict(col_map)
                col_map["cvc"], col_map["__expiry__"] = col_map["__expiry__"], col_map["cvc"]
                stats["expiry_cvc_swapped"] += 1

        # 姓名可能是分开两列，也可能是单列「持卡人」这种
        first, last = cell("first_name"), cell("last_name")
        if not (first and last):
            full_first, full_last = split_name(cell("__fullname__"))
            first, last = first or full_first, last or full_last

        out.append(make_card(
            number, month, year, cell("cvc"), first, last,
            [cell("country"), cell("__country2__")],
            cell("address"), cell("address2"), cell("city"),
            cell("state"), cell("zip"),
        ))
    return out


def validate(card, today=None, require_address=False, excluded=()):
    """返回剔除原因；None 表示通过。"""
    needed = REQUIRED_FIELDS + (ADDRESS_REQUIRED if require_address else [])
    missing = [f for f in needed if not card.get(f)]
    if missing:
        return f"缺字段: {','.join(missing)}"
    if card["country"] and card["country"] in excluded:
        return f"国家黑名单({card['country']})"
    n = card["card_number"]
    if not 13 <= len(n) <= 19:
        return f"卡号长度异常({len(n)})"
    if not luhn_ok(n):
        return "Luhn 校验失败"
    if len(card["cvc"]) not in (3, 4):
        return f"CVC 长度异常({len(card['cvc'])})"
    if today is not None:
        exp = (int(card["expiry_year"]), int(card["expiry_month"]))
        if exp < today:
            return "已过期"
        # 12/72 这类笔误会解析成 2072 年，真卡不会有这么长的有效期
        if exp[0] > today[0] + 15:
            return "有效期年份离谱"
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default="底料")
    ap.add_argument("--out", default="底料/合并结果.xlsx")
    ap.add_argument("--keep-expired", action="store_true", help="保留已过期的卡（默认剔除）")
    ap.add_argument("--require-address", action="store_true",
                    help="只保留账单信息齐全的卡（默认地址缺失也收）")
    ap.add_argument("--exclude-countries", default=",".join(DEFAULT_EXCLUDED_COUNTRIES),
                    help="剔除这些国家的卡，逗号分隔，MX / Mexico / 墨西哥 三种写法都认，"
                         f"默认 {','.join(DEFAULT_EXCLUDED_COUNTRIES)}")
    ap.add_argument("--no-exclude-countries", action="store_true",
                    help="不按国家剔除，所有国家全收")
    args = ap.parse_args()

    from datetime import date
    today = None if args.keep_expired else (date.today().year, date.today().month)

    excluded = set()
    if not args.no_exclude_countries:
        for name in args.exclude_countries.split(","):
            name = name.strip()
            if not name:
                continue
            code = normalize_country(name)
            if not code:
                print(f"--exclude-countries 认不出这个国家: {name}")
                return 1
            excluded.add(code)

    files = sorted(
        os.path.join(args.src, f)
        for f in os.listdir(args.src)
        if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")
    )
    files = [f for f in files if os.path.abspath(f) != os.path.abspath(args.out)]
    if not files:
        print(f"{args.src} 下没有找到表格文件")
        return 1

    print(f"国家黑名单    {','.join(sorted(excluded)) if excluded else '（未启用）'}")

    stats = Counter()
    merged = {}          # card_number -> card（同卡号保留账单信息最全的那条）
    reasons = Counter()

    for path in files:
        try:
            wb = load_workbook_any(path)
        except Exception as e:
            print(f"  !! {os.path.basename(path)}: {e}")
            stats["file_failed"] += 1
            continue

        file_new = 0
        file_raw = 0
        for ws in wb.worksheets:
            for card in extract_rows(ws, stats):
                file_raw += 1
                stats["raw"] += 1
                reason = validate(card, today, args.require_address, excluded)
                if reason:
                    reasons[reason.split("(")[0].split(":")[0]] += 1
                    stats["invalid"] += 1
                    continue
                key = card["card_number"]
                seen = merged.get(key)
                if seen is not None:
                    stats["dup"] += 1
                    # 同一张卡在几十个表里重复出现，各表填的账单信息详略不一，
                    # 保留最全的那条，否则会被先扫到的残缺记录顶掉
                    if completeness(card) > completeness(seen):
                        merged[key] = card
                        stats["dup_upgraded"] += 1
                    continue
                merged[key] = card
                file_new += 1
        wb.close()
        print(f"  {os.path.basename(path):<32} 原始 {file_raw:>5} 行 → 新增 {file_new:>4} 张")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CreditCards"
    ws.append(TEMPLATE_COLUMNS)
    # 账单信息齐全的排前面，下游先用好数据
    cards = sorted(merged.values(), key=completeness, reverse=True)
    # 卡号/邮编等必须是文本，否则 Excel 会吃掉前导零、长数字转科学计数
    for card in cards:
        ws.append([card[c] for c in TEMPLATE_COLUMNS])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.number_format = "@"
    wb.save(args.out)

    full = sum(1 for c in cards if all(c[f] for f in ADDRESS_REQUIRED))
    print()
    print(f"扫描文件      {len(files)}（失败 {stats['file_failed']}，跳过无关 sheet {stats['sheet_skipped']}）")
    print(f"原始卡行      {stats['raw']}"
          f"（其中 JSON 串里捞出 {stats['json_card']}，解析失败 {stats['json_unparsable']}）")
    print(f"剔除无效      {stats['invalid']}  " + ", ".join(f"{k}×{v}" for k, v in reasons.most_common()))
    print(f"重复卡号      {stats['dup']}（其中 {stats['dup_upgraded']} 次用更全的记录替换了先前的）")
    print(f"输出          {args.out}  共 {len(cards)} 张"
          f"（账单信息齐全 {full}，不齐 {len(cards) - full}）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
